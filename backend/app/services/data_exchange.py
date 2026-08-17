"""Shared, schema-driven data exchange for institution-configured College data."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, false, func, or_, select
from sqlalchemy.orm import Session

from app.models import (
    Client,
    CollegeAssessment,
    CollegeAssessmentComponent,
    CollegeAssessmentScore,
    CollegeAssessmentScheme,
    CollegeAttendanceSnapshot,
    CollegeCareerEvidence,
    CollegeClearanceSnapshot,
    CollegeCohort,
    CollegeCourse,
    CollegeCourseOffering,
    CollegeDepartment,
    CollegeExamCycle,
    CollegeImportRun,
    CollegeProgram,
    CollegeStudentProfile,
    CollegeTerm,
    CollegeTermResult,
    CollegeApplicationStageEvent,
    CollegeCodingAccount,
    CollegePipelineStage,
    CollegePlacementApplication,
    CollegePlacementCompany,
    CollegePlacementOpportunity,
    CollegeReadinessSnapshot,
    CollegePlacementInterview,
    CollegePlacementOffer,
    AuditLog,
    DataExchangeArtifact,
    DataExchangeRow,
    DataExchangeRun,
)
from app.services.college_assessments import (
    build_scheme_snapshot, component_payload, freeze_scheme, normalize_code,
    recalculate_assessment_score, validate_component_definitions, validate_metric_values,
)
from app.services.college_imports import RESOURCE_FIELDS, commit_run, stage_rows, validate_row


MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 10_000
MAX_EXPORT_ROWS = 5_000
MAX_WORKSHEETS = 20
MAX_COLUMNS = 200
CLEAR_SENTINEL = "__CLEAR__"
STRUCTURE_ORDER = ("departments", "programs", "cohorts", "terms", "courses", "offerings")
STRUCTURE_PACKAGE_ORDER = (*STRUCTURE_ORDER, "assessment_schemes")
STRUCTURE_SHEET_NAMES = {
    "departments": "Departments",
    "programs": "Programs",
    "cohorts": "Cohorts",
    "terms": "Terms",
    "courses": "Courses",
    "offerings": "Offerings",
    "assessment_schemes": "Assessment Patterns",
}
STRUCTURE_MODELS = {
    "departments": CollegeDepartment,
    "programs": CollegeProgram,
    "cohorts": CollegeCohort,
    "terms": CollegeTerm,
    "courses": CollegeCourse,
    "offerings": CollegeCourseOffering,
}


@dataclass(frozen=True)
class ExchangeField:
    key: str
    label: str
    field_type: str = "text"
    required: bool = False
    writable: bool = True
    max_length: int | None = None
    lookup: str | None = None
    description: str | None = None

    def payload(self) -> dict:
        return {
            "key": self.key,
            "label": self.label,
            "type": self.field_type,
            "required": self.required,
            "writable": self.writable,
            "max_length": self.max_length,
            "lookup": self.lookup,
            "description": self.description,
        }


@dataclass(frozen=True)
class ExchangeResource:
    key: str
    label: str
    category: str
    fields: tuple[ExchangeField, ...] = field(default_factory=tuple)
    methods: tuple[str, ...] = ("manual", "excel", "csv", "erp_pull", "api_push")
    importable: bool = True
    exportable: bool = True
    update_supported: bool = False
    permission: str = "college.imports.manage"
    description: str = ""


def _f(key: str, label: str, field_type="text", required=False, **kwargs) -> ExchangeField:
    return ExchangeField(key, label, field_type, required, **kwargs)


RESOURCES: dict[str, ExchangeResource] = {
    "academic_structure": ExchangeResource(
        "academic_structure", "Academic structure package", "Academic structure",
        methods=("excel", "csv"), update_supported=True,
        description="Departments, programs, cohorts, terms, courses, offerings, and assessment patterns in dependency order.",
    ),
    "departments": ExchangeResource("departments", "Departments", "Academic structure", (
        _f("external_id", "External ID"), _f("name", "Department name", required=True),
        _f("code", "Department code", required=True), _f("description", "Description"),
    ), update_supported=True),
    "programs": ExchangeResource("programs", "Programs", "Academic structure", (
        _f("external_id", "External ID"), _f("department_code", "Department code", required=True, lookup="departments"),
        _f("name", "Program name", required=True), _f("code", "Program code", required=True),
        _f("degree_type", "Degree type", required=True), _f("duration_semesters", "Duration in terms", "integer", True),
    ), update_supported=True),
    "cohorts": ExchangeResource("cohorts", "Graduation batches and sections", "Academic structure", (
        _f("external_id", "External ID"), _f("program_code", "Program code", required=True, lookup="programs"),
        _f("name", "Cohort name", required=True), _f("code", "Cohort code", required=True),
        _f("admission_year", "Admission year", "integer", True), _f("graduation_year", "Graduation year", "integer", True),
        _f("current_semester", "Current term", "integer", True), _f("section", "Section", required=True),
    ), update_supported=True),
    "terms": ExchangeResource("terms", "Academic years and terms", "Academic structure", (
        _f("external_id", "External ID"), _f("name", "Term name", required=True),
        _f("academic_year", "Academic year", required=True), _f("term_number", "Term number", "integer", True),
        _f("starts_on", "Starts on", "date", True), _f("ends_on", "Ends on", "date", True),
        _f("status", "Status", required=True), _f("is_current", "Current term", "boolean"),
    ), update_supported=True),
    "courses": ExchangeResource("courses", "Courses", "Academic structure", (
        _f("external_id", "External ID"), _f("department_code", "Department code", required=True, lookup="departments"),
        _f("name", "Course name", required=True), _f("code", "Course code", required=True),
        _f("credits", "Credits", "integer", True), _f("course_type", "Course type", required=True),
    ), update_supported=True),
    "offerings": ExchangeResource("offerings", "Course offerings", "Academic structure", (
        _f("term_id", "Term ID", required=True, lookup="terms"), _f("course_id", "Course ID", required=True, lookup="courses"),
        _f("cohort_id", "Cohort ID", required=True, lookup="cohorts"), _f("room", "Room"),
        _f("status", "Status", required=True),
    ), methods=("manual", "excel", "csv"), update_supported=True),
    "assessment_schemes": ExchangeResource(
        "assessment_schemes", "Assessment patterns", "Assessments", methods=("manual", "excel", "csv"),
        description="Versioned component definitions. Used versions remain immutable.",
    ),
    "exam_cycles": ExchangeResource(
        "exam_cycles", "Exam cycles", "Assessments", methods=("manual", "excel", "csv", "erp_pull", "api_push"),
        description="Exam or test cycles created from an effective assessment pattern.",
    ),
    "assessment_marks": ExchangeResource(
        "assessment_marks", "Marks and custom metrics", "Assessments", methods=("manual", "excel", "csv", "erp_pull", "api_push"),
        update_supported=True,
        description="Dynamic columns are generated from the selected exam cycle's frozen pattern version.",
    ),
    "students": ExchangeResource("students", "Students", "Students", (
        _f("record_id", "Student ID"), _f("version", "Version", "integer"),
        *tuple(
        _f(key, key.replace("_", " ").title(), required=key in {"admission_number", "first_name", "program_code", "cohort_code"})
        for key in RESOURCE_FIELDS["students"]
        ),
    ), update_supported=True),
    "term_results": ExchangeResource("term_results", "Term results", "Academic evidence", (
        _f("record_id", "Result ID"), _f("version", "Version", "integer"),
        *tuple(
        _f(key, key.replace("_", " ").title(), required=key in {"admission_number", "semester"})
        for key in RESOURCE_FIELDS["term_results"]
        ),
    ), update_supported=True),
    "attendance": ExchangeResource("attendance", "Attendance snapshots", "Academic evidence", (
        _f("record_id", "Attendance ID"), _f("version", "Version", "integer"),
        *tuple(
        _f(key, key.replace("_", " ").title(), required=key in {"admission_number", "as_of"})
        for key in RESOURCE_FIELDS["attendance"]
        ),
    ), update_supported=True),
    "skills": ExchangeResource("skills", "Career evidence", "Student enrichment", (
        _f("record_id", "Evidence ID"), _f("version", "Version", "integer"),
        *tuple(
        _f(key, key.replace("_", " ").title(), required=key in {"admission_number", "title"})
        for key in RESOURCE_FIELDS["skills"]
        ),
    ), update_supported=True),
    "coding_accounts": ExchangeResource(
        "coding_accounts", "Coding accounts", "Student enrichment", (
            _f("admission_number", "Admission number", required=True),
            _f("platform", "Platform", required=True), _f("username", "Username", required=True),
            _f("consent_status", "Consent status", required=True),
            _f("verification_status", "Verification status", required=True),
        ), methods=("manual", "excel", "csv"), permission="college.coding.manage",
    ),
    "companies": ExchangeResource("companies", "Recruiting companies", "Placements", (
        _f("name", "Company name", required=True), _f("industry", "Industry"),
        _f("website", "Website"), _f("contact_name", "Contact name"),
        _f("contact_email", "Contact email"), _f("contact_phone", "Contact phone"),
    ), methods=("manual", "excel", "csv"), permission="college.companies.manage"),
    "drives": ExchangeResource("drives", "Placement drives", "Placements", (
        _f("company_id", "Company ID", required=True, lookup="companies"),
        _f("title", "Drive title", required=True), _f("opportunity_type", "Opportunity type", required=True),
        _f("status", "Status", required=True), _f("opens_at", "Opens at", "datetime"),
        _f("deadline_at", "Deadline", "datetime"), _f("drive_at", "Drive date", "datetime"),
        _f("work_location", "Work location"), _f("employment_type", "Employment type"),
        _f("package_min_paise", "Minimum package (paise)", "integer"),
        _f("package_max_paise", "Maximum package (paise)", "integer"),
    ), methods=("manual", "excel", "csv"), permission="college.opportunities.manage"),
    "applications": ExchangeResource(
        "applications", "Applications and stage updates", "Placements", (
            _f("record_id", "Application ID", required=True), _f("version", "Version", "integer", True),
            _f("stage_code", "Stage code", required=True), _f("reason", "Reason", required=True),
        ), methods=("manual", "excel", "csv"), update_supported=True,
        permission="college.applications.manage",
        description="Stage changes use record IDs and versions and create normal stage history.",
    ),
    "internship_clearance": ExchangeResource(
        "internship_clearance", "Internship clearance", "Restricted", tuple(
            _f(key, key.replace("_", " ").title(), required=key in {"admission_number", "status", "as_of"})
            for key in RESOURCE_FIELDS["internship_clearance"]
        ), methods=("erp_pull", "api_push"), permission="college.clearance.view",
        description="ERP-authoritative clearance status only; fee amounts are never exchanged.",
    ),
    "readiness": ExchangeResource("readiness", "Readiness snapshots", "Exports", (
        _f("admission_number", "Admission number", writable=False), _f("student_name", "Student name", writable=False),
        _f("score", "Readiness score", "number", writable=False), _f("coverage_percent", "Evidence coverage", "percentage", writable=False),
        _f("band", "Readiness band", writable=False), _f("calculated_at", "Calculated at", "datetime", writable=False),
    ), methods=(), importable=False, permission="college.readiness.view"),
    "leaderboards": ExchangeResource("leaderboards", "Leaderboards", "Exports", (
        _f("rank", "Rank", "rank", writable=False), _f("admission_number", "Admission number", writable=False),
        _f("student_name", "Student name", writable=False), _f("score", "Score", "number", writable=False),
        _f("coverage_percent", "Evidence coverage", "percentage", writable=False), _f("band", "Band", writable=False),
    ), methods=(), importable=False, permission="college.readiness.view"),
    "eligibility": ExchangeResource("eligibility", "Eligibility evidence", "Exports", (
        _f("application_id", "Application ID", writable=False), _f("admission_number", "Admission number", writable=False),
        _f("opportunity", "Opportunity", writable=False), _f("eligibility_status", "Eligibility status", writable=False),
        _f("evidence", "Evidence snapshot", writable=False), _f("evaluated_at", "Evaluated at", "datetime", writable=False),
    ), methods=(), importable=False, permission="college.placements.view"),
    "interviews": ExchangeResource("interviews", "Interviews", "Exports", (
        _f("application_id", "Application ID", writable=False), _f("admission_number", "Admission number", writable=False),
        _f("interview_type", "Interview type", writable=False), _f("scheduled_at", "Scheduled at", "datetime", writable=False),
        _f("status", "Status", writable=False), _f("score_percent", "Score", "percentage", writable=False),
    ), methods=(), importable=False, permission="college.placements.view"),
    "offers": ExchangeResource("offers", "Offers", "Exports", (
        _f("application_id", "Application ID", writable=False), _f("admission_number", "Admission number", writable=False),
        _f("offered_role", "Offered role", writable=False), _f("package_paise", "Package (paise)", "integer", writable=False),
        _f("offered_on", "Offered on", "date", writable=False), _f("joining_on", "Joining on", "date", writable=False),
        _f("status", "Status", writable=False),
    ), methods=(), importable=False, permission="college.placements.view"),
    "outcomes": ExchangeResource("outcomes", "Placement outcomes", "Exports", (
        _f("application_id", "Application ID", writable=False), _f("admission_number", "Admission number", writable=False),
        _f("opportunity", "Opportunity", writable=False), _f("company", "Company", writable=False),
        _f("outcome", "Outcome", writable=False), _f("current_stage", "Current stage", writable=False),
    ), methods=(), importable=False, permission="college.placements.view"),
    "audits": ExchangeResource("audits", "Audit history", "Exports", (
        _f("event", "Event", writable=False), _f("resource_type", "Resource type", writable=False),
        _f("resource_id", "Resource ID", writable=False), _f("actor_id", "Actor ID", writable=False),
        _f("rows_affected", "Rows affected", "integer", writable=False), _f("created_at", "Time", "datetime", writable=False),
    ), methods=(), importable=False, permission="college.academics.view"),
}


RESOURCE_DOMAINS = {
    "academic_structure": "academics", "departments": "academics", "programs": "academics",
    "cohorts": "academics", "terms": "academics", "courses": "academics", "offerings": "academics",
    "assessment_schemes": "academics", "exam_cycles": "assessments", "assessment_marks": "assessments",
    "students": "students", "term_results": "assessments", "attendance": "attendance",
    "skills": "coding", "coding_accounts": "coding", "companies": "placements",
    "drives": "placements", "applications": "placements", "internship_clearance": "clearance",
    "readiness": "readiness", "leaderboards": "readiness", "eligibility": "placements",
    "interviews": "placements", "offers": "placements", "outcomes": "placements", "audits": "reports",
}

RESOURCE_VIEW_PERMISSIONS = {
    "academic_structure": "college.academics.view", "departments": "college.academics.view",
    "programs": "college.academics.view", "cohorts": "college.academics.view",
    "terms": "college.academics.view", "courses": "college.academics.view",
    "offerings": "college.academics.view", "assessment_schemes": "college.academics.view",
    "exam_cycles": "college.assessments.view", "assessment_marks": "college.assessments.view",
    "students": "college.students.view", "term_results": "college.assessments.view",
    "attendance": "college.attendance.view", "skills": "college.coding.view",
    "coding_accounts": "college.coding.view", "companies": "college.placements.view",
    "drives": "college.placements.view", "applications": "college.placements.view",
    "internship_clearance": "college.clearance.view", "readiness": "college.readiness.view",
    "leaderboards": "college.readiness.view", "eligibility": "college.placements.view",
    "interviews": "college.placements.view", "offers": "college.placements.view",
    "outcomes": "college.placements.view", "audits": "audit.view",
}

RESOURCE_WRITE_PERMISSIONS = {
    "academic_structure": "college.academics.manage", "departments": "college.academics.manage",
    "programs": "college.academics.manage", "cohorts": "college.academics.manage",
    "terms": "college.academics.manage", "courses": "college.academics.manage",
    "offerings": "college.academics.manage", "assessment_schemes": "college.academics.manage",
    "exam_cycles": "college.assessments.manage", "assessment_marks": "college.assessments.record",
    "students": "college.students.manage", "term_results": "college.assessments.record",
    "attendance": "college.attendance.mark", "skills": "college.coding.manage",
    "coding_accounts": "college.coding.manage", "companies": "college.companies.manage",
    "drives": "college.opportunities.manage", "applications": "college.applications.manage",
    "internship_clearance": "college.clearance.manage",
}


def _can_import_resource(resource: ExchangeResource, permission_codes: set[str]) -> bool:
    if not resource.importable or "college.imports.manage" not in permission_codes:
        return False
    required = RESOURCE_WRITE_PERMISSIONS.get(resource.key)
    return bool(required and required in permission_codes)


def resource_catalog(permission_codes: set[str]) -> list[dict]:
    result = []
    for resource in RESOURCES.values():
        if RESOURCE_VIEW_PERMISSIONS.get(resource.key) not in permission_codes:
            continue
        can_import = _can_import_resource(resource, permission_codes)
        result.append({
            "key": resource.key,
            "label": resource.label,
            "category": resource.category,
            "description": resource.description,
            "methods": list(resource.methods) if can_import else [],
            "importable": can_import,
            "exportable": resource.exportable and "college.data.export" in permission_codes,
            "update_supported": resource.update_supported if can_import else False,
        })
    return result


def _cycle(db: Session, organization_id: str, scope: dict) -> CollegeExamCycle:
    cycle_id = str(scope.get("cycle_id") or "")
    if not cycle_id:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Select an exam cycle before generating marks fields")
    cycle = db.scalar(select(CollegeExamCycle).where(
        CollegeExamCycle.id == cycle_id,
        CollegeExamCycle.organization_id == organization_id,
    ))
    if not cycle:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Exam cycle not found")
    return cycle


def resource_schema(db: Session, organization_id: str, resource_key: str, scope: dict | None = None) -> dict:
    resource = RESOURCES.get(resource_key)
    if not resource:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Data exchange resource not found")
    scope = dict(scope or {})
    fields = list(resource.fields)
    effective = None
    if resource_key == "assessment_marks":
        cycle = _cycle(db, organization_id, scope)
        components = list(cycle.scheme_snapshot.get("components") or [])
        if cycle.scheme_component_id:
            components = [item for item in components if item.get("id") == cycle.scheme_component_id]
        fields = [
            _f("assessment_id", "Assessment ID", required=True),
            _f("record_id", "Score record ID"),
            _f("version", "Record version", "integer"),
            _f("admission_number", "Admission number", required=True),
            _f("student_name", "Student name", writable=False),
        ] + [ExchangeField(
            key=str(item["code"]),
            label=str(item["name"]),
            field_type=str(item.get("metric_type") or "number"),
            required=bool(item.get("is_required")),
            description=(f"Maximum {item['max_marks']}" if item.get("max_marks") is not None else None),
        ) for item in components]
        effective = {
            "cycle_id": cycle.id,
            "cycle_code": cycle.code,
            "scheme_code": cycle.scheme_snapshot.get("scheme_code"),
            "scheme_version": cycle.scheme_snapshot.get("scheme_version"),
            "calculation_method": cycle.scheme_snapshot.get("calculation_method"),
        }
        assessments = list(db.scalars(select(CollegeAssessment).where(
            CollegeAssessment.organization_id == organization_id,
            CollegeAssessment.exam_cycle_id == cycle.id,
        )))
        allowed_cohort_ids = scope.get("allowed_cohort_ids")
        allowed_offering_ids = scope.get("allowed_course_offering_ids")
        if allowed_cohort_ids is not None or allowed_offering_ids is not None:
            allowed_cohorts = set(allowed_cohort_ids or [])
            allowed_offerings = set(allowed_offering_ids or [])
            assessments = [item for item in assessments if (
                (item.cohort_id and item.cohort_id in allowed_cohorts)
                or (item.offering_id and item.offering_id in allowed_offerings)
            )]
            if not assessments:
                raise HTTPException(status.HTTP_404_NOT_FOUND, "Assessment cycle not found")
        cohort_ids = {item.cohort_id for item in assessments if item.cohort_id}
        if any(item.offering_id and not item.cohort_id for item in assessments):
            offering_ids = [item.offering_id for item in assessments if item.offering_id and not item.cohort_id]
            cohort_ids.update(db.scalars(select(CollegeCourseOffering.cohort_id).where(
                CollegeCourseOffering.organization_id == organization_id,
                CollegeCourseOffering.id.in_(offering_ids),
            )))
        cohorts = db.execute(select(CollegeCohort, CollegeProgram).join(
            CollegeProgram, CollegeProgram.id == CollegeCohort.program_id,
        ).where(
            CollegeCohort.organization_id == organization_id,
            CollegeCohort.id.in_(cohort_ids),
        ).order_by(CollegeProgram.code, CollegeCohort.graduation_year, CollegeCohort.section)).all() if cohort_ids else []
        effective["available_cohorts"] = [{
            "id": cohort.id,
            "code": cohort.code,
            "name": cohort.name,
            "section": cohort.section,
            "program_code": program.code,
            "graduation_year": cohort.graduation_year,
        } for cohort, program in cohorts]
    elif resource_key == "assessment_schemes":
        fields = [
            _f("scheme_code", "Pattern code", required=True), _f("scheme_version", "Pattern version", "integer", True),
            _f("scheme_name", "Pattern name", required=True), _f("domain", "Domain", required=True),
            _f("calculation_method", "Calculation method", required=True), _f("final_score_max", "Final score scale", "number", True),
            _f("component_code", "Component code", required=True), _f("component_name", "Component name", required=True),
            _f("metric_type", "Metric type", required=True), _f("max_marks", "Maximum", "number"),
            _f("weightage_percent", "Weightage %", "percentage"), _f("pass_marks", "Pass threshold", "number"),
            _f("required", "Required", "boolean"), _f("best_n", "Best N", "integer"),
            _f("minimum_components", "Minimum components", "integer"),
        ]
    elif resource_key == "exam_cycles":
        fields = [
            _f("scheme_code", "Pattern code", required=True), _f("scheme_version", "Pattern version", "integer", True),
            _f("component_code", "Component code"), _f("cycle_code", "Cycle code", required=True),
            _f("cycle_name", "Cycle name", required=True), _f("term_id", "Term ID", lookup="terms"),
            _f("held_on", "Held on", "date"), _f("due_on", "Due on", "date"),
            _f("offering_ids", "Offering IDs", description="Comma-separated for academic cycles"),
            _f("cohort_ids", "Cohort IDs", description="Comma-separated for coding or placement cycles"),
        ]
    elif resource_key == "applications":
        fields = [_f("record_id", "Application ID", required=True), _f("version", "Version", "integer", True), _f("stage_code", "Stage code", required=True), _f("reason", "Reason", required=True)]
    elif resource_key == "offerings":
        fields = list(resource.fields)

    return {
        "resource": {
            "key": resource.key, "label": resource.label, "category": resource.category,
            "description": resource.description, "methods": list(resource.methods),
            "importable": resource.importable, "exportable": resource.exportable,
            "update_supported": resource.update_supported,
        },
        "schema_version": "1",
        "fields": [item.payload() for item in fields],
        "effective_configuration": effective,
        "clear_sentinel": CLEAR_SENTINEL,
        "limits": {"file_bytes": MAX_FILE_BYTES, "rows": MAX_ROWS, "worksheets": MAX_WORKSHEETS, "columns": MAX_COLUMNS},
        "push_payload": {
            "scheme_code": effective.get("scheme_code") if effective else None,
            "scheme_version": effective.get("scheme_version") if effective else None,
            "cycle_code": effective.get("cycle_code") if effective else None,
            "records": [{
                "student": {"admission_number": "college admission number"},
                "academic_scope": {"assessment_id": "required when a cohort has multiple assessments in this cycle"},
                "metrics": {field.key: "value" for field in fields if field.key not in {"assessment_id", "record_id", "version", "admission_number", "student_name"}},
            }],
        } if resource_key == "assessment_marks" else None,
    }


def lookup_rows(db: Session, organization_id: str, scope: dict | None = None) -> dict[str, list[dict]]:
    scope = scope or {}
    departments = list(db.scalars(select(CollegeDepartment).where(CollegeDepartment.organization_id == organization_id).order_by(CollegeDepartment.code)))
    programs = list(db.scalars(select(CollegeProgram).where(CollegeProgram.organization_id == organization_id).order_by(CollegeProgram.code)))
    cohorts = list(db.scalars(select(CollegeCohort).where(CollegeCohort.organization_id == organization_id).order_by(CollegeCohort.code)))
    terms = list(db.scalars(select(CollegeTerm).where(CollegeTerm.organization_id == organization_id).order_by(CollegeTerm.starts_on.desc())))
    courses = list(db.scalars(select(CollegeCourse).where(CollegeCourse.organization_id == organization_id).order_by(CollegeCourse.code)))
    offerings = list(db.scalars(select(CollegeCourseOffering).where(CollegeCourseOffering.organization_id == organization_id).order_by(CollegeCourseOffering.created_at.desc()).limit(5000)))
    companies = list(db.scalars(select(CollegePlacementCompany).where(
        CollegePlacementCompany.organization_id == organization_id,
        CollegePlacementCompany.is_active.is_(True),
    ).order_by(CollegePlacementCompany.name)))
    stages = list(db.scalars(select(CollegePipelineStage).where(
        CollegePipelineStage.organization_id == organization_id,
        CollegePipelineStage.is_enabled.is_(True),
    ).order_by(CollegePipelineStage.display_order)))
    allowed_departments = scope.get("allowed_department_ids")
    allowed_programs = scope.get("allowed_program_ids")
    allowed_cohorts = scope.get("allowed_cohort_ids")
    allowed_offerings = scope.get("allowed_course_offering_ids")
    if allowed_departments is not None:
        departments = [row for row in departments if row.id in set(allowed_departments)]
        courses = [row for row in courses if row.department_id in set(allowed_departments)]
    if allowed_programs is not None:
        programs = [row for row in programs if row.id in set(allowed_programs)]
    if allowed_cohorts is not None:
        cohorts = [row for row in cohorts if row.id in set(allowed_cohorts)]
    if allowed_offerings is not None:
        offerings = [row for row in offerings if row.id in set(allowed_offerings)]
    return {
        "departments": [{"id": row.id, "code": row.code, "name": row.name} for row in departments],
        "programs": [{"id": row.id, "code": row.code, "name": row.name, "department_id": row.department_id} for row in programs],
        "cohorts": [{"id": row.id, "code": row.code, "name": row.name, "program_id": row.program_id, "section": row.section} for row in cohorts],
        "terms": [{"id": row.id, "code": f"{row.academic_year}-{row.term_number}", "name": row.name} for row in terms],
        "courses": [{"id": row.id, "code": row.code, "name": row.name, "department_id": row.department_id} for row in courses],
        "offerings": [{"id": row.id, "term_id": row.term_id, "course_id": row.course_id, "cohort_id": row.cohort_id} for row in offerings],
        "companies": [{"id": row.id, "code": row.id, "name": row.name} for row in companies],
        "pipeline_stages": [{"id": row.id, "code": row.slug, "name": row.name} for row in stages],
    }


def _structure_rows(db: Session, organization_id: str, key: str, *, include_identity: bool = True) -> list[dict]:
    if key == "departments":
        rows = list(db.scalars(select(CollegeDepartment).where(CollegeDepartment.organization_id == organization_id).order_by(CollegeDepartment.code)))
        result = [{"name": row.name, "code": row.code, "description": row.description, "record_id": row.id, "version": row.version} for row in rows]
    elif key == "programs":
        rows = db.execute(select(CollegeProgram, CollegeDepartment).join(CollegeDepartment, CollegeDepartment.id == CollegeProgram.department_id).where(CollegeProgram.organization_id == organization_id).order_by(CollegeProgram.code)).all()
        result = [{"department_code": department.code, "name": row.name, "code": row.code, "degree_type": row.degree_type, "duration_semesters": row.duration_semesters, "record_id": row.id, "version": row.version} for row, department in rows]
    elif key == "cohorts":
        rows = db.execute(select(CollegeCohort, CollegeProgram).join(CollegeProgram, CollegeProgram.id == CollegeCohort.program_id).where(CollegeCohort.organization_id == organization_id).order_by(CollegeCohort.code)).all()
        result = [{"program_code": program.code, "name": row.name, "code": row.code, "admission_year": row.admission_year, "graduation_year": row.graduation_year, "current_semester": row.current_semester, "section": row.section, "record_id": row.id, "version": row.version} for row, program in rows]
    elif key == "terms":
        rows = list(db.scalars(select(CollegeTerm).where(CollegeTerm.organization_id == organization_id).order_by(CollegeTerm.starts_on)))
        result = [{"name": row.name, "academic_year": row.academic_year, "term_number": row.term_number, "starts_on": row.starts_on, "ends_on": row.ends_on, "status": row.status, "is_current": row.is_current, "record_id": row.id, "version": row.version} for row in rows]
    elif key == "courses":
        rows = db.execute(select(CollegeCourse, CollegeDepartment).join(CollegeDepartment, CollegeDepartment.id == CollegeCourse.department_id).where(CollegeCourse.organization_id == organization_id).order_by(CollegeCourse.code)).all()
        result = [{"department_code": department.code, "name": row.name, "code": row.code, "credits": row.credits, "course_type": row.course_type, "record_id": row.id, "version": row.version} for row, department in rows]
    elif key == "offerings":
        rows = list(db.scalars(select(CollegeCourseOffering).where(CollegeCourseOffering.organization_id == organization_id).order_by(CollegeCourseOffering.created_at)))
        result = [{"term_id": row.term_id, "course_id": row.course_id, "cohort_id": row.cohort_id, "room": row.room, "status": row.status, "record_id": row.id, "version": row.version} for row in rows]
    else:
        result = []
    if not include_identity:
        for item in result:
            item.pop("record_id", None)
            item.pop("version", None)
    return result


def marks_template_rows(db: Session, organization_id: str, scope: dict) -> list[dict]:
    cycle = _cycle(db, organization_id, scope)
    assessments = list(db.scalars(select(CollegeAssessment).where(
        CollegeAssessment.organization_id == organization_id,
        CollegeAssessment.exam_cycle_id == cycle.id,
    )))
    selected_assessment_ids = {str(value) for value in scope.get("assessment_ids") or []}
    selected_cohort_ids = {str(value) for value in scope.get("cohort_ids") or []}
    if selected_assessment_ids:
        assessments = [item for item in assessments if item.id in selected_assessment_ids]
    result: list[dict] = []
    for assessment in assessments:
        cohort_id = assessment.cohort_id
        if not cohort_id and assessment.offering_id:
            cohort_id = db.scalar(select(CollegeCourseOffering.cohort_id).where(CollegeCourseOffering.id == assessment.offering_id))
        if not cohort_id:
            continue
        if selected_cohort_ids and cohort_id not in selected_cohort_ids:
            continue
        rows_statement = select(CollegeStudentProfile, Client, CollegeAssessmentScore).join(
            Client, Client.id == CollegeStudentProfile.client_id,
        ).outerjoin(CollegeAssessmentScore, and_(
            CollegeAssessmentScore.assessment_id == assessment.id,
            CollegeAssessmentScore.student_profile_id == CollegeStudentProfile.id,
        )).where(
            CollegeStudentProfile.organization_id == organization_id,
            CollegeStudentProfile.cohort_id == cohort_id,
            CollegeStudentProfile.status == "active",
        )
        allowed_student_ids = scope.get("allowed_student_ids")
        if allowed_student_ids is not None:
            rows_statement = rows_statement.where(
                CollegeStudentProfile.id.in_(allowed_student_ids) if allowed_student_ids else false()
            )
        rows = db.execute(rows_statement.order_by(
            CollegeStudentProfile.roll_number, CollegeStudentProfile.admission_number,
        )).all()
        for profile, client, score in rows:
            item = {
                "assessment_id": assessment.id,
                "record_id": score.id if score else None,
                "version": score.version if score else None,
                "admission_number": profile.admission_number,
                "student_name": f"{client.first_name} {client.last_name}".strip(),
            }
            if score:
                item.update(score.metrics or {})
                if len(assessment.metric_schema or []) == 1 and score.marks_awarded is not None:
                    item.setdefault(str(assessment.metric_schema[0]["code"]), float(score.marks_awarded))
            result.append(item)
            if len(result) > MAX_ROWS:
                raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "This cycle exceeds the 10,000-row workbook limit; narrow the scope")
    return result


def template_rows(db: Session, organization_id: str, resource_key: str, mode: str, scope: dict) -> list[dict]:
    if resource_key in STRUCTURE_MODELS:
        return _structure_rows(db, organization_id, resource_key) if mode == "update" else []
    if resource_key == "assessment_marks":
        return marks_template_rows(db, organization_id, scope)
    if resource_key == "applications" and mode == "update":
        statement = select(CollegePlacementApplication, CollegeStudentProfile).join(
            CollegeStudentProfile,
            CollegeStudentProfile.id == CollegePlacementApplication.student_profile_id,
        ).where(CollegePlacementApplication.organization_id == organization_id)
        allowed_student_ids = scope.get("allowed_student_ids")
        if allowed_student_ids is not None:
            statement = statement.where(
                CollegeStudentProfile.id.in_(allowed_student_ids) if allowed_student_ids else false()
            )
        rows = db.execute(statement.order_by(
            CollegePlacementApplication.updated_at.desc(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "version": row.version,
            "admission_number": student.admission_number,
            "stage_code": "", "reason": "",
        } for row, student in rows]
    return []


def _export_rows_unfiltered(db: Session, organization_id: str, resource_key: str, scope: dict) -> list[dict]:
    """Return bounded, tenant-scoped source rows for synchronous exports."""
    if resource_key in STRUCTURE_MODELS:
        return _structure_rows(db, organization_id, resource_key)
    if resource_key == "assessment_marks":
        return marks_template_rows(db, organization_id, scope)
    if resource_key == "students":
        rows = db.execute(select(CollegeStudentProfile, Client, CollegeProgram, CollegeCohort).join(
            Client, Client.id == CollegeStudentProfile.client_id,
        ).join(CollegeProgram, CollegeProgram.id == CollegeStudentProfile.program_id).join(
            CollegeCohort, CollegeCohort.id == CollegeStudentProfile.cohort_id,
        ).where(CollegeStudentProfile.organization_id == organization_id).order_by(
            CollegeStudentProfile.admission_number,
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": profile.id, "version": profile.version,
            "admission_number": profile.admission_number, "first_name": client.first_name,
            "last_name": client.last_name, "email": client.email, "phone": client.phone,
            "program_code": program.code, "cohort_code": cohort.code,
            "current_semester": profile.current_semester,
        } for profile, client, program, cohort in rows]
    if resource_key == "term_results":
        rows = db.execute(select(CollegeTermResult, CollegeStudentProfile).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeTermResult.student_profile_id,
        ).where(CollegeTermResult.organization_id == organization_id).order_by(
            CollegeTermResult.created_at.desc(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "version": row.version, "admission_number": student.admission_number,
            "semester": row.semester, "sgpa": row.sgpa, "cgpa": row.cgpa,
            "active_backlogs": row.active_backlogs, "total_backlogs": row.total_backlogs,
            "credits_earned": row.credits_earned, "published_on": row.published_on,
        } for row, student in rows]
    if resource_key == "attendance":
        rows = db.execute(select(CollegeAttendanceSnapshot, CollegeStudentProfile).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeAttendanceSnapshot.student_profile_id,
        ).where(CollegeAttendanceSnapshot.organization_id == organization_id).order_by(
            CollegeAttendanceSnapshot.as_of.desc(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "version": row.version, "admission_number": student.admission_number,
            "scope": row.scope_key, "classes_held": row.classes_held,
            "classes_attended": row.classes_attended,
            "attendance_percent": row.attendance_percent, "as_of": row.as_of,
        } for row, student in rows]
    if resource_key == "skills":
        rows = db.execute(select(CollegeCareerEvidence, CollegeStudentProfile).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeCareerEvidence.student_profile_id,
        ).where(
            CollegeCareerEvidence.organization_id == organization_id,
            CollegeCareerEvidence.evidence_type == "skill",
        ).order_by(
            CollegeStudentProfile.admission_number, CollegeCareerEvidence.title,
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "version": row.version, "external_id": row.external_id,
            "admission_number": student.admission_number, "title": row.title,
            "proficiency": row.proficiency, "verified": row.is_verified,
            "evidence_url": row.evidence_url,
        } for row, student in rows]
    if resource_key == "internship_clearance":
        rows = db.execute(select(CollegeClearanceSnapshot, CollegeStudentProfile).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeClearanceSnapshot.student_profile_id,
        ).where(CollegeClearanceSnapshot.organization_id == organization_id).order_by(
            CollegeClearanceSnapshot.as_of.desc(), CollegeStudentProfile.admission_number,
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "external_id": row.external_id, "admission_number": student.admission_number,
            "status": row.status, "as_of": row.as_of,
            "source_updated_at": row.source_updated_at,
        } for row, student in rows]
    if resource_key == "assessment_schemes":
        schemes = list(db.scalars(select(CollegeAssessmentScheme).where(
            CollegeAssessmentScheme.organization_id == organization_id,
        ).order_by(CollegeAssessmentScheme.code, CollegeAssessmentScheme.version_number)))
        result = []
        for scheme in schemes:
            components = list(db.scalars(select(CollegeAssessmentComponent).where(
                CollegeAssessmentComponent.scheme_id == scheme.id,
            ).order_by(CollegeAssessmentComponent.display_order)))
            for component in components:
                result.append({
                    "record_id": scheme.id,
                    "scheme_code": scheme.code, "scheme_version": scheme.version_number,
                    "scheme_name": scheme.name, "domain": scheme.domain,
                    "calculation_method": scheme.calculation_method,
                    "final_score_max": scheme.final_score_max,
                    "component_code": component.code, "component_name": component.name,
                    "metric_type": component.metric_type, "max_marks": component.max_marks,
                    "weightage_percent": component.weightage_bps / 100,
                    "pass_marks": component.pass_marks, "required": component.is_required,
                    "best_n": (scheme.calculation_config or {}).get("best_n"),
                    "minimum_components": (scheme.calculation_config or {}).get("minimum_components"),
                })
        return result
    if resource_key == "exam_cycles":
        rows = list(db.scalars(select(CollegeExamCycle).where(
            CollegeExamCycle.organization_id == organization_id,
        ).order_by(CollegeExamCycle.created_at.desc()).limit(MAX_ROWS + 1)))
        return [{
            "record_id": row.id, "version": row.version,
            "scheme_code": (row.scheme_snapshot or {}).get("scheme_code"),
            "scheme_version": (row.scheme_snapshot or {}).get("scheme_version"),
            "component_code": next((item.get("code") for item in (row.scheme_snapshot or {}).get("components", []) if item.get("id") == row.scheme_component_id), None),
            "cycle_code": row.code, "cycle_name": row.name, "term_id": row.term_id,
            "held_on": row.held_on, "due_on": row.due_on,
            "offering_ids": ",".join(row.target_offering_ids or []),
            "cohort_ids": ",".join(row.target_cohort_ids or []),
        } for row in rows]
    if resource_key == "coding_accounts":
        rows = db.execute(select(CollegeCodingAccount, CollegeStudentProfile).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeCodingAccount.student_profile_id,
        ).where(CollegeCodingAccount.organization_id == organization_id).order_by(
            CollegeStudentProfile.admission_number, CollegeCodingAccount.platform,
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "admission_number": student.admission_number, "platform": row.platform,
            "username": row.username, "consent_status": row.consent_status,
            "verification_status": row.verification_status,
        } for row, student in rows]
    if resource_key == "companies":
        rows = list(db.scalars(select(CollegePlacementCompany).where(
            CollegePlacementCompany.organization_id == organization_id,
        ).order_by(CollegePlacementCompany.name).limit(MAX_ROWS + 1)))
        return [{
            "record_id": row.id, "name": row.name, "industry": row.industry, "website": row.website,
            "contact_name": row.contact_name, "contact_email": row.contact_email,
            "contact_phone": row.contact_phone,
        } for row in rows]
    if resource_key == "drives":
        rows = list(db.scalars(select(CollegePlacementOpportunity).where(
            CollegePlacementOpportunity.organization_id == organization_id,
        ).order_by(CollegePlacementOpportunity.created_at.desc()).limit(MAX_ROWS + 1)))
        return [{
            "record_id": row.id, "company_id": row.company_id, "title": row.title, "opportunity_type": row.opportunity_type,
            "status": row.status, "opens_at": row.opens_at, "deadline_at": row.deadline_at,
            "drive_at": row.drive_at, "work_location": row.work_location,
            "employment_type": row.employment_type, "package_min_paise": row.package_min_paise,
            "package_max_paise": row.package_max_paise,
        } for row in rows]
    if resource_key == "applications":
        rows = db.execute(select(CollegePlacementApplication, CollegePipelineStage).outerjoin(
            CollegePipelineStage, CollegePipelineStage.id == CollegePlacementApplication.current_stage_id,
        ).where(CollegePlacementApplication.organization_id == organization_id).order_by(
            CollegePlacementApplication.updated_at.desc(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "version": row.version,
            "stage_code": stage.slug if stage else None, "reason": None,
        } for row, stage in rows]
    if resource_key in {"readiness", "leaderboards"}:
        rows = db.execute(select(CollegeReadinessSnapshot, CollegeStudentProfile, Client).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegeReadinessSnapshot.student_profile_id,
        ).join(Client, Client.id == CollegeStudentProfile.client_id).where(
            CollegeReadinessSnapshot.organization_id == organization_id,
        ).order_by(
            CollegeReadinessSnapshot.score.desc().nullslast(), CollegeStudentProfile.admission_number,
        ).limit(MAX_ROWS + 1)).all()
        output = [{
            "record_id": row.id,
            "admission_number": student.admission_number,
            "student_name": f"{client.first_name} {client.last_name}".strip(),
            "score": row.score, "coverage_percent": row.coverage_percent,
            "band": row.band, "calculated_at": row.calculated_at,
        } for row, student, client in rows]
        if resource_key == "leaderboards":
            return [{"rank": index, **{key: value for key, value in item.items() if key != "calculated_at"}} for index, item in enumerate(output, start=1)]
        return output
    if resource_key == "eligibility":
        rows = db.execute(select(
            CollegePlacementApplication, CollegeStudentProfile, CollegePlacementOpportunity,
        ).join(CollegeStudentProfile, CollegeStudentProfile.id == CollegePlacementApplication.student_profile_id).join(
            CollegePlacementOpportunity, CollegePlacementOpportunity.id == CollegePlacementApplication.opportunity_id,
        ).where(CollegePlacementApplication.organization_id == organization_id).order_by(
            CollegePlacementApplication.eligibility_evaluated_at.desc().nullslast(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "application_id": row.id, "admission_number": student.admission_number,
            "opportunity": opportunity.title, "eligibility_status": row.eligibility_status,
            "evidence": json.dumps(row.eligibility_evidence or {}, sort_keys=True, default=str),
            "evaluated_at": row.eligibility_evaluated_at,
        } for row, student, opportunity in rows]
    if resource_key == "interviews":
        rows = db.execute(select(
            CollegePlacementInterview, CollegePlacementApplication, CollegeStudentProfile,
        ).join(CollegePlacementApplication, CollegePlacementApplication.id == CollegePlacementInterview.application_id).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegePlacementApplication.student_profile_id,
        ).where(CollegePlacementInterview.organization_id == organization_id).order_by(
            CollegePlacementInterview.scheduled_at.desc().nullslast(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "application_id": application.id, "admission_number": student.admission_number,
            "interview_type": row.interview_type, "scheduled_at": row.scheduled_at,
            "status": row.status, "score_percent": row.score_percent,
        } for row, application, student in rows]
    if resource_key == "offers":
        rows = db.execute(select(
            CollegePlacementOffer, CollegePlacementApplication, CollegeStudentProfile,
        ).join(CollegePlacementApplication, CollegePlacementApplication.id == CollegePlacementOffer.application_id).join(
            CollegeStudentProfile, CollegeStudentProfile.id == CollegePlacementApplication.student_profile_id,
        ).where(CollegePlacementOffer.organization_id == organization_id).order_by(
            CollegePlacementOffer.offered_on.desc().nullslast(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "application_id": application.id, "admission_number": student.admission_number,
            "offered_role": row.offered_role, "package_paise": row.package_paise,
            "offered_on": row.offered_on, "joining_on": row.joining_on, "status": row.status,
        } for row, application, student in rows]
    if resource_key == "outcomes":
        rows = db.execute(select(
            CollegePlacementApplication, CollegeStudentProfile, CollegePlacementOpportunity,
            CollegePlacementCompany, CollegePipelineStage,
        ).join(CollegeStudentProfile, CollegeStudentProfile.id == CollegePlacementApplication.student_profile_id).join(
            CollegePlacementOpportunity, CollegePlacementOpportunity.id == CollegePlacementApplication.opportunity_id,
        ).join(CollegePlacementCompany, CollegePlacementCompany.id == CollegePlacementOpportunity.company_id).outerjoin(
            CollegePipelineStage, CollegePipelineStage.id == CollegePlacementApplication.current_stage_id,
        ).where(CollegePlacementApplication.organization_id == organization_id).order_by(
            CollegePlacementApplication.updated_at.desc(),
        ).limit(MAX_ROWS + 1)).all()
        return [{
            "record_id": row.id, "application_id": row.id, "admission_number": student.admission_number,
            "opportunity": opportunity.title, "company": company.name, "outcome": row.outcome,
            "current_stage": stage.name if stage else None,
        } for row, student, opportunity, company, stage in rows]
    if resource_key == "audits":
        rows = list(db.scalars(select(AuditLog).where(
            AuditLog.organization_id == organization_id,
            or_(
                AuditLog.action.like("college.%"),
                AuditLog.action.like("data_exchange.%"),
            ),
        ).order_by(AuditLog.created_at.desc()).limit(MAX_ROWS + 1)))
        return [{
            "record_id": row.id, "event": row.action, "resource_type": row.resource_type,
            "resource_id": row.resource_id, "actor_id": row.user_id,
            "rows_affected": row.rows_affected, "created_at": row.created_at,
        } for row in rows]
    return []


def export_rows(db: Session, organization_id: str, resource_key: str, scope: dict) -> list[dict]:
    rows = _export_rows_unfiltered(db, organization_id, resource_key, scope)
    allowed_student_ids = scope.get("allowed_student_ids")
    if allowed_student_ids is not None:
        allowed_students = list(db.scalars(select(CollegeStudentProfile).where(
            CollegeStudentProfile.organization_id == organization_id,
            CollegeStudentProfile.id.in_(allowed_student_ids) if allowed_student_ids else false(),
        )))
        allowed_admissions = {row.admission_number for row in allowed_students}
        if resource_key == "students":
            rows = [row for row in rows if str(row.get("record_id")) in set(allowed_student_ids)]
        elif rows and any("admission_number" in row for row in rows):
            rows = [row for row in rows if row.get("admission_number") in allowed_admissions]

    direct_scope = {
        "departments": ("allowed_department_ids", "record_id"),
        "programs": ("allowed_program_ids", "record_id"),
        "cohorts": ("allowed_cohort_ids", "record_id"),
        "offerings": ("allowed_course_offering_ids", "record_id"),
    }.get(resource_key)
    if direct_scope and scope.get(direct_scope[0]) is not None:
        allowed = {str(value) for value in scope.get(direct_scope[0]) or []}
        rows = [row for row in rows if str(row.get(direct_scope[1])) in allowed]
    if scope.get("selection") != "selected":
        return rows
    selected_ids = {str(value) for value in scope.get("selected_ids") or []}
    return [
        row for row in rows
        if str(row.get("record_id") or row.get("application_id") or "") in selected_ids
    ]


def estimate_export_rows(db: Session, organization_id: str, resource_key: str, scope: dict) -> int:
    """Estimate export size cheaply so only genuinely large files use a worker."""
    selected_ids = list(dict.fromkeys(scope.get("selected_ids") or []))
    if scope.get("selection") == "selected":
        return len(selected_ids)
    if resource_key == "academic_structure":
        return sum(
            int(db.scalar(select(func.count()).select_from(model).where(model.organization_id == organization_id)) or 0)
            for model in STRUCTURE_MODELS.values()
        )
    if resource_key == "assessment_marks":
        cycle = _cycle(db, organization_id, scope)
        assessments = list(db.scalars(select(CollegeAssessment).where(
            CollegeAssessment.organization_id == organization_id,
            CollegeAssessment.exam_cycle_id == cycle.id,
        )))
        selected_assessment_ids = {str(value) for value in scope.get("assessment_ids") or []}
        selected_cohort_ids = {str(value) for value in scope.get("cohort_ids") or []}
        if selected_assessment_ids:
            assessments = [item for item in assessments if item.id in selected_assessment_ids]
        total = 0
        for assessment in assessments:
            cohort_id = assessment.cohort_id
            if not cohort_id and assessment.offering_id:
                cohort_id = db.scalar(select(CollegeCourseOffering.cohort_id).where(
                    CollegeCourseOffering.id == assessment.offering_id,
                ))
            if cohort_id:
                if selected_cohort_ids and cohort_id not in selected_cohort_ids:
                    continue
                total += int(db.scalar(select(func.count()).select_from(CollegeStudentProfile).where(
                    CollegeStudentProfile.organization_id == organization_id,
                    CollegeStudentProfile.cohort_id == cohort_id,
                    CollegeStudentProfile.status == "active",
                )) or 0)
        return total
    model_by_resource = {
        **STRUCTURE_MODELS,
        "students": CollegeStudentProfile,
        "term_results": CollegeTermResult,
        "attendance": CollegeAttendanceSnapshot,
        "assessment_schemes": CollegeAssessmentComponent,
        "exam_cycles": CollegeExamCycle,
        "skills": CollegeCareerEvidence,
        "coding_accounts": CollegeCodingAccount,
        "companies": CollegePlacementCompany,
        "drives": CollegePlacementOpportunity,
        "applications": CollegePlacementApplication,
        "internship_clearance": CollegeClearanceSnapshot,
        "readiness": CollegeReadinessSnapshot,
        "leaderboards": CollegeReadinessSnapshot,
        "eligibility": CollegePlacementApplication,
        "interviews": CollegePlacementInterview,
        "offers": CollegePlacementOffer,
        "outcomes": CollegePlacementApplication,
        "audits": AuditLog,
    }
    model = model_by_resource.get(resource_key)
    if not model:
        return 0
    statement = select(func.count()).select_from(model).where(model.organization_id == organization_id)
    if resource_key == "skills":
        statement = statement.where(CollegeCareerEvidence.evidence_type == "skill")
    elif resource_key == "audits":
        statement = statement.where(or_(
            AuditLog.action.like("college.%"), AuditLog.action.like("data_exchange.%"),
        ))
    return int(db.scalar(statement) or 0)


def _columns(schema: dict, *, mode: str) -> list[str]:
    columns = [field["key"] for field in schema["fields"]]
    if mode == "update" and schema["resource"]["update_supported"]:
        columns = ["record_id", "version", *[item for item in columns if item not in {"record_id", "version"}]]
    return list(dict.fromkeys(columns))


def create_csv_template(schema: dict, rows: list[dict], *, mode: str) -> bytes:
    output = io.StringIO(newline="")
    columns = _columns(schema, mode=mode)
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key) for key in columns})
    return output.getvalue().encode("utf-8-sig")


def _format_sheet(sheet, columns: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="124A38")
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(34, max(14, len(column) + 3))


def create_excel_template(
    db: Session,
    organization_id: str,
    schema: dict,
    rows: list[dict],
    *,
    mode: str,
    scope: dict,
) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append([schema["resource"]["label"]])
    instructions.append(["Mode", mode])
    instructions.append(["How to use", "Enter values only in the Data sheet. Keep IDs and versions unchanged in update files."])
    instructions.append(["Blank update cells", "Leave the stored value unchanged."])
    instructions.append(["Clear optional value", f"Enter {CLEAR_SENTINEL} exactly."])
    instructions.append(["Safety", "Uploads are previewed first. Nothing is committed until you approve the valid rows."])
    instructions["A1"].font = Font(bold=True, size=16, color="124A38")
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 90

    data = workbook.create_sheet("Data")
    columns = _columns(schema, mode=mode)
    data.append(columns)
    for row in rows:
        data.append([row.get(column) for column in columns])
    _format_sheet(data, columns)

    lookups = workbook.create_sheet("Lookups")
    lookup_data = lookup_rows(db, organization_id, scope)
    lookup_columns = ["resource", "id", "code", "name", "parent_id", "section"]
    lookups.append(lookup_columns)
    for resource_key, items in lookup_data.items():
        for item in items:
            lookups.append([
                resource_key, item.get("id"), item.get("code"), item.get("name"),
                item.get("department_id") or item.get("program_id"), item.get("section"),
            ])
    _format_sheet(lookups, lookup_columns)

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["resource_key", schema["resource"]["key"]])
    metadata.append(["schema_version", schema["schema_version"]])
    metadata.append(["mode", mode])
    metadata.append(["scope", json.dumps(scope, separators=(",", ":"), sort_keys=True)])
    metadata.append(["effective_configuration", json.dumps(schema.get("effective_configuration"), separators=(",", ":"), sort_keys=True)])
    metadata.sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _lookup_csv(db: Session, organization_id: str) -> bytes:
    output = io.StringIO(newline="")
    columns = ["resource", "id", "code", "name", "parent_id", "section"]
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for resource_key, items in lookup_rows(db, organization_id).items():
        for item in items:
            writer.writerow({
                "resource": resource_key,
                "id": item.get("id"),
                "code": item.get("code"),
                "name": item.get("name"),
                "parent_id": item.get("department_id") or item.get("program_id"),
                "section": item.get("section"),
            })
    return output.getvalue().encode("utf-8-sig")


def create_structure_package(
    db: Session,
    organization_id: str,
    file_format: str,
    mode: str,
    *,
    include_scheme_rows: bool = False,
) -> tuple[bytes, str, str]:
    if file_format == "csv":
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
            manifest = []
            for order, key in enumerate(STRUCTURE_PACKAGE_ORDER, start=1):
                schema = resource_schema(db, organization_id, key)
                rows = (
                    _export_rows_unfiltered(db, organization_id, key, {})
                    if key == "assessment_schemes" and include_scheme_rows
                    else template_rows(db, organization_id, key, mode, {})
                )
                filename = f"{order:02d}_{key}.csv"
                archive.writestr(filename, create_csv_template(schema, rows, mode=mode))
                manifest.append({"order": order, "resource": key, "file": filename})
            archive.writestr("lookups.csv", _lookup_csv(db, organization_id))
            archive.writestr("manifest.json", json.dumps({
                "dependency_order": manifest,
                "lookups": "lookups.csv",
                "instructions": "Import one CSV at a time in dependency order. Assessment patterns are last and never replace an existing pattern implicitly.",
            }, indent=2))
        return output.getvalue(), "academic-structure.zip", "application/zip"

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Academic structure package"])
    instructions.append(["Dependency order", "Departments, Programs, Cohorts, Terms, Courses, Offerings, Assessment Patterns"])
    instructions.append(["Blank update cells", "Stored values remain unchanged"])
    instructions.append(["Clear optional value", CLEAR_SENTINEL])
    instructions.append(["Assessment patterns", "Each component is one row. Existing patterns are immutable and are never replaced by an upload."])
    instructions["A1"].font = Font(bold=True, size=16, color="124A38")
    for key in STRUCTURE_PACKAGE_ORDER:
        schema = resource_schema(db, organization_id, key)
        columns = _columns(schema, mode=mode)
        sheet = workbook.create_sheet(STRUCTURE_SHEET_NAMES[key])
        sheet.append(columns)
        rows = (
            _export_rows_unfiltered(db, organization_id, key, {})
            if key == "assessment_schemes" and include_scheme_rows
            else template_rows(db, organization_id, key, mode, {})
        )
        for row in rows:
            sheet.append([row.get(column) for column in columns])
        _format_sheet(sheet, columns)
    lookups = workbook.create_sheet("Lookups")
    lookups.append(["resource", "id", "code", "name", "parent_id", "section"])
    for resource_key, items in lookup_rows(db, organization_id).items():
        for item in items:
            lookups.append([resource_key, item.get("id"), item.get("code"), item.get("name"), item.get("department_id") or item.get("program_id"), item.get("section")])
    _format_sheet(lookups, ["resource", "id", "code", "name", "parent_id", "section"])
    metadata = workbook.create_sheet("Metadata")
    metadata.append(["resource_key", "academic_structure"])
    metadata.append(["schema_version", "1"])
    metadata.append(["mode", mode])
    metadata.sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), "academic-structure.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def generate_template(
    db: Session,
    organization_id: str,
    resource_key: str,
    *,
    file_format: str,
    mode: str,
    scope: dict,
) -> tuple[bytes, str, str]:
    if resource_key == "academic_structure":
        return create_structure_package(db, organization_id, file_format, mode)
    schema = resource_schema(db, organization_id, resource_key, scope)
    rows = template_rows(db, organization_id, resource_key, mode, scope)
    if file_format == "csv":
        return create_csv_template(schema, rows, mode=mode), f"{resource_key}-{mode}.csv", "text/csv"
    return (
        create_excel_template(db, organization_id, schema, rows, mode=mode, scope=scope),
        f"{resource_key}-{mode}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def generate_export(
    db: Session,
    organization_id: str,
    resource_key: str,
    *,
    file_format: str,
    scope: dict,
) -> tuple[bytes, str, str, int]:
    if resource_key == "academic_structure":
        content, filename, content_type = create_structure_package(
            db, organization_id, file_format, "update", include_scheme_rows=True,
        )
        return content, filename.replace("academic-structure", "academic-structure-export"), content_type, estimate_export_rows(
            db, organization_id, resource_key, scope,
        )
    schema = resource_schema(db, organization_id, resource_key, scope)
    rows = export_rows(db, organization_id, resource_key, scope)
    if len(rows) > MAX_ROWS:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "This export exceeds the 10,000-row file limit; narrow the filters")
    if file_format == "csv":
        content = create_csv_template(schema, rows, mode="update")
        return content, f"{resource_key}-export.csv", "text/csv", len(rows)
    content = create_excel_template(db, organization_id, schema, rows, mode="update", scope=scope)
    return content, f"{resource_key}-export.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", len(rows)


def parse_upload(content: bytes, filename: str) -> tuple[list[dict], dict]:
    if not content:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The uploaded file is empty")
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "Files must be 10 MB or smaller")
    lowered = filename.casefold()
    if lowered.endswith((".xlsm", ".xltm", ".xls", ".ods")):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Upload a macro-free .xlsx or .csv file")
    if lowered.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "CSV files must use UTF-8 encoding") from exc
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The CSV header row is missing")
        if len(reader.fieldnames) > MAX_COLUMNS:
            raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "Files may contain at most 200 columns")
        rows = [{str(key).strip(): value for key, value in row.items() if key} for row in reader]
        if len(rows) > MAX_ROWS:
            raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "Files may contain at most 10,000 data rows")
        return rows, {}
    if not lowered.endswith(".xlsx"):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Upload a .csv or macro-free .xlsx file")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=False, keep_vba=False)
    except Exception as exc:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The Excel workbook could not be read") from exc
    if len(workbook.sheetnames) > MAX_WORKSHEETS:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "Workbooks may contain at most 20 worksheets")
    metadata = {}
    if "Metadata" in workbook.sheetnames:
        metadata = {str(row[0].value): row[1].value for row in workbook["Metadata"].iter_rows(min_row=1, max_col=2) if row[0].value}
    sheets = [workbook["Data"]] if "Data" in workbook.sheetnames else [
        workbook[name] for name in workbook.sheetnames if name not in {"Instructions", "Lookups", "Metadata"}
    ]
    all_rows: list[dict] = []
    for sheet in sheets:
        if sheet.max_column > MAX_COLUMNS:
            raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, f"{sheet.title} has more than 200 columns")
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        if not any(headers):
            continue
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2, max_col=len(headers)), start=2):
            if any(cell.data_type == "f" for cell in cells):
                raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, f"Formulas are not allowed ({sheet.title}, row {row_number})")
            values = {header: cell.value for header, cell in zip(headers, cells) if header}
            if not any(value not in (None, "") for value in values.values()):
                continue
            if len(sheets) > 1:
                values["_resource_key"] = sheet.title.strip().casefold().replace(" ", "_")
            all_rows.append(values)
            if len(all_rows) > MAX_ROWS:
                raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "Workbooks may contain at most 10,000 data rows")
    return all_rows, metadata


def store_artifact(
    db: Session,
    run: DataExchangeRun,
    *,
    kind: str,
    filename: str,
    content_type: str,
    content: bytes,
) -> DataExchangeArtifact:
    existing = db.scalar(select(DataExchangeArtifact).where(
        DataExchangeArtifact.run_id == run.id,
        DataExchangeArtifact.kind == kind,
    ))
    if existing:
        db.delete(existing)
        db.flush()
    artifact = DataExchangeArtifact(
        organization_id=run.organization_id,
        run_id=run.id,
        kind=kind,
        filename=filename,
        content_type=content_type,
        byte_size=len(content),
        checksum=hashlib.sha256(content).hexdigest(),
        content=content,
    )
    db.add(artifact)
    db.flush()
    return artifact


def run_payload(run: DataExchangeRun) -> dict:
    return {
        "id": run.id,
        "resource_key": run.resource_key,
        "operation": run.operation,
        "source_type": run.source_type,
        "file_format": run.file_format,
        "status": run.status,
        "scope": run.scope,
        "row_count": run.row_count,
        "valid_count": run.valid_count,
        "create_count": run.create_count,
        "update_count": run.update_count,
        "unchanged_count": run.unchanged_count,
        "invalid_count": run.invalid_count,
        "conflict_count": run.conflict_count,
        "committed_count": run.committed_count,
        "source_filename": run.source_filename,
        "created_at": run.created_at,
        "completed_at": run.completed_at,
        "artifacts": {
            "template": f"/api/data-exchange/runs/{run.id}/artifacts/template" if run.operation == "template" else None,
            "export": f"/api/data-exchange/runs/{run.id}/artifacts/export" if run.operation == "export" else None,
            "corrections": f"/api/data-exchange/runs/{run.id}/artifacts/corrections" if run.invalid_count else None,
        },
    }


def _structure_current(db: Session, organization_id: str, key: str, record_id: str) -> tuple[Any | None, dict]:
    model = STRUCTURE_MODELS[key]
    record = db.scalar(select(model).where(model.id == record_id, model.organization_id == organization_id))
    if not record:
        return None, {}
    rows = _structure_rows(db, organization_id, key)
    current = next((item for item in rows if item.get("record_id") == record_id), {})
    return record, current


def _stage_structure_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict, key: str) -> DataExchangeRow:
    record_id = str(raw.get("record_id") or "").strip() or None
    version_value = raw.get("version")
    record = None
    current = {}
    errors: list[str] = []
    parsed_version = _whole_number(version_value, "version", errors, minimum=1)
    action = "create"
    if record_id:
        record, current = _structure_current(db, run.organization_id, key, record_id)
        if not record:
            errors.append("record_id was not found in this college")
        elif version_value in (None, ""):
            errors.append("version is required for updates")
        elif parsed_version is not None and parsed_version != record.version:
            errors.append("record version is stale; download a new update template")
        action = "update"
    values = {}
    source_fields = [field for field in RESOURCE_FIELDS.get(key, ()) if field not in {"external_id", "source_updated_at"}]
    if key == "offerings":
        source_fields = ["term_id", "course_id", "cohort_id", "room", "status"]
    for field_name in source_fields:
        incoming = raw.get(field_name)
        if action == "update" and incoming in (None, ""):
            values[field_name] = current.get(field_name)
        elif incoming == CLEAR_SENTINEL:
            if field_name not in {"description", "room"}:
                errors.append(f"{field_name} cannot be cleared")
            values[field_name] = None
        else:
            values[field_name] = incoming
    if key in RESOURCE_FIELDS:
        normalized, validation_errors = validate_row(values, key)
        values = normalized
        errors.extend(validation_errors)
    else:
        for required in ("term_id", "course_id", "cohort_id"):
            if not str(values.get(required) or "").strip():
                errors.append(f"{required} is required")
        if values.get("status") not in {"active", "inactive", "closed"}:
            errors.append("status must be active, inactive, or closed")
    changes = {key: value for key, value in values.items() if key not in {"external_id", "source_updated_at"} and current.get(key) != value}
    if action == "update" and not changes and not errors:
        action = "unchanged"
    status_value = "invalid" if errors else "valid"
    return DataExchangeRow(
        organization_id=run.organization_id,
        run_id=run.id,
        row_number=row_number,
        action=action,
        status=status_value,
        natural_key=str(values.get("code") or values.get("name") or record_id or row_number),
        record_id=record_id,
        record_version=parsed_version,
        values={**values, "_resource_key": key},
        current_values=current,
        changes=changes,
        errors=errors,
    )


def _stage_marks_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    cycle = _cycle(db, run.organization_id, run.scope)
    assessment_id = str(raw.get("assessment_id") or "").strip()
    admission_number = str(raw.get("admission_number") or "").strip()
    errors = []
    assessment = db.scalar(select(CollegeAssessment).where(
        CollegeAssessment.id == assessment_id,
        CollegeAssessment.organization_id == run.organization_id,
        CollegeAssessment.exam_cycle_id == cycle.id,
    )) if assessment_id else None
    if not assessment:
        errors.append("assessment_id is not part of the selected cycle")
    selected_assessment_ids = {str(value) for value in (run.scope or {}).get("assessment_ids") or []}
    selected_cohort_ids = {str(value) for value in (run.scope or {}).get("cohort_ids") or []}
    if assessment and selected_assessment_ids and assessment.id not in selected_assessment_ids:
        errors.append("assessment_id is outside the selected workbook scope")
    student = db.scalar(select(CollegeStudentProfile).where(
        CollegeStudentProfile.organization_id == run.organization_id,
        CollegeStudentProfile.admission_number == admission_number,
    )) if admission_number else None
    if not student:
        errors.append("admission_number was not found")
    allowed_student_ids = (run.mapping or {}).get("allowed_student_ids")
    if student and allowed_student_ids is not None and student.id not in set(allowed_student_ids):
        errors.append("student is outside your College access")
    elif assessment and student.cohort_id != assessment.cohort_id:
        errors.append("student does not belong to this assessment cohort")
    if student and selected_cohort_ids and student.cohort_id not in selected_cohort_ids:
        errors.append("student is outside the selected workbook cohort scope")
    record_id = str(raw.get("record_id") or "").strip() or None
    score = db.scalar(select(CollegeAssessmentScore).where(
        CollegeAssessmentScore.id == record_id,
        CollegeAssessmentScore.organization_id == run.organization_id,
        CollegeAssessmentScore.assessment_id == assessment_id,
    )) if record_id else None
    existing = db.scalar(select(CollegeAssessmentScore).where(
        CollegeAssessmentScore.organization_id == run.organization_id,
        CollegeAssessmentScore.assessment_id == assessment_id,
        CollegeAssessmentScore.student_profile_id == student.id,
    )) if assessment and student else None
    action = "update" if record_id else "create"
    if record_id and not score:
        errors.append("score record_id was not found")
    if not record_id and existing:
        errors.append("A score already exists; download an update template instead of implicitly replacing it")
    version_value = raw.get("version")
    parsed_version = _whole_number(version_value, "version", errors, minimum=1)
    if score and version_value in (None, ""):
        errors.append("version is required for score updates")
    elif score and parsed_version is not None and parsed_version != score.version:
        errors.append("score version is stale; download a new update template")
    definitions = list(assessment.metric_schema or []) if assessment else []
    known_metric_codes = {str(item["code"]) for item in definitions}
    reserved_columns = {"assessment_id", "record_id", "version", "admission_number", "student_name", "_row_number"}
    unknown_columns = sorted(
        str(key) for key, value in raw.items()
        if key not in reserved_columns and key not in known_metric_codes and value not in (None, "")
    )
    if unknown_columns:
        errors.append(f"Unknown metric columns: {', '.join(unknown_columns[:10])}")
    current_metrics = dict(score.metrics or {}) if score else {}
    if score and len(definitions) == 1 and score.marks_awarded is not None:
        current_metrics.setdefault(str(definitions[0]["code"]), float(score.marks_awarded))
    metrics = {}
    for definition in definitions:
        code = str(definition["code"])
        incoming = raw.get(code)
        if score and incoming in (None, ""):
            if code in current_metrics:
                metrics[code] = current_metrics[code]
        elif incoming == CLEAR_SENTINEL:
            if definition.get("is_required"):
                errors.append(f"{definition['name']} is required and cannot be cleared")
        elif incoming not in (None, ""):
            metrics[code] = incoming
    try:
        normalized_metrics = validate_metric_values(definitions, metrics, allow_partial=False)
    except ValueError as exc:
        normalized_metrics = metrics
        errors.append(str(exc))
    changes = {key: value for key, value in normalized_metrics.items() if current_metrics.get(key) != value}
    if action == "update" and not changes and not errors:
        action = "unchanged"
    return DataExchangeRow(
        organization_id=run.organization_id,
        run_id=run.id,
        row_number=row_number,
        action=action,
        status="invalid" if errors else "valid",
        natural_key=f"{assessment_id}:{admission_number}",
        record_id=record_id,
        record_version=parsed_version,
        values={"assessment_id": assessment_id, "student_profile_id": student.id if student else None, "metrics": normalized_metrics},
        current_values={"metrics": current_metrics},
        changes={"metrics": changes},
        errors=errors,
    )


def _clean_text(value: Any, *, maximum: int | None = None) -> str | None:
    if value in (None, "", CLEAR_SENTINEL):
        return None
    result = " ".join(str(value).split())
    return result[:maximum] if maximum else result


def _whole_number(value: Any, field_name: str, errors: list[str], *, minimum: int = 0) -> int | None:
    if value in (None, ""):
        return None
    try:
        parsed = Decimal(str(value))
        if not parsed.is_finite() or parsed != parsed.to_integral_value():
            raise InvalidOperation
        number = int(parsed)
    except (InvalidOperation, TypeError, ValueError):
        errors.append(f"{field_name} must be a whole number")
        return None
    if number < minimum:
        errors.append(f"{field_name} must be at least {minimum}")
    return number


def _number_value(
    value: Any,
    field_name: str,
    errors: list[str],
    *,
    minimum: Decimal | None = None,
    maximum: Decimal | None = None,
    required: bool = False,
) -> float | None:
    if value in (None, ""):
        if required:
            errors.append(f"{field_name} is required")
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        errors.append(f"{field_name} must be a number")
        return None
    if not number.is_finite():
        errors.append(f"{field_name} must be finite")
        return None
    if minimum is not None and number < minimum:
        errors.append(f"{field_name} must be at least {minimum}")
    if maximum is not None and number > maximum:
        errors.append(f"{field_name} must be at most {maximum}")
    return float(number)


def _exchange_code(value: Any, field_name: str, errors: list[str], *, max_length: int) -> str:
    if value in (None, ""):
        return ""
    try:
        return normalize_code(str(value), max_length=max_length)
    except HTTPException:
        errors.append(f"{field_name} must contain at least one letter or number")
        return ""


def _boolean_value(value: Any, field_name: str, errors: list[str], *, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"true", "yes", "y", "1", "required"}:
        return True
    if normalized in {"false", "no", "n", "0", "optional"}:
        return False
    errors.append(f"{field_name} must be true or false")
    return default


def _date_or_datetime(value: Any, field_name: str, errors: list[str], *, with_time: bool = False):
    if value in (None, "", CLEAR_SENTINEL):
        return None
    if isinstance(value, datetime):
        return value if with_time else value.date()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc) if with_time else value
    try:
        text_value = str(value).strip().replace("Z", "+00:00")
        return datetime.fromisoformat(text_value) if with_time else date.fromisoformat(text_value[:10])
    except ValueError:
        errors.append(f"{field_name} must use a valid {'date and time' if with_time else 'date'}")
        return None


def _base_exchange_row(run: DataExchangeRun, row_number: int, *, action: str, values: dict, errors: list[str], natural_key: str, record_id: str | None = None, record_version: int | None = None, current_values: dict | None = None, changes: dict | None = None) -> DataExchangeRow:
    return DataExchangeRow(
        organization_id=run.organization_id,
        run_id=run.id,
        row_number=row_number,
        action=action,
        status="invalid" if errors else "valid",
        natural_key=natural_key,
        record_id=record_id,
        record_version=record_version,
        values=values,
        current_values=current_values or {},
        changes=changes or values,
        errors=errors,
    )


def _stage_scheme_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    scheme_code = _exchange_code(raw.get("scheme_code"), "scheme_code", errors, max_length=50)
    component_code = _exchange_code(raw.get("component_code"), "component_code", errors, max_length=50)
    scheme_version = _whole_number(raw.get("scheme_version"), "scheme_version", errors, minimum=1)
    name = _clean_text(raw.get("scheme_name"), maximum=180)
    component_name = _clean_text(raw.get("component_name"), maximum=140)
    domain = str(raw.get("domain") or "").strip().casefold()
    method = str(raw.get("calculation_method") or "").strip().casefold()
    metric_type = str(raw.get("metric_type") or "number").strip().casefold()
    final_score_max = _number_value(
        raw.get("final_score_max"), "final_score_max", errors,
        minimum=Decimal("0.01"), maximum=Decimal("1000000"), required=True,
    )
    max_marks = _number_value(
        raw.get("max_marks"), "max_marks", errors,
        minimum=Decimal("0.01"), maximum=Decimal("1000000"),
    )
    weightage_percent = _number_value(
        raw.get("weightage_percent"), "weightage_percent", errors,
        minimum=Decimal("0"), maximum=Decimal("100"),
    )
    pass_marks = _number_value(
        raw.get("pass_marks"), "pass_marks", errors,
        minimum=Decimal("0"), maximum=Decimal("1000000"),
    )
    best_n = _whole_number(raw.get("best_n"), "best_n", errors, minimum=1)
    minimum_components = _whole_number(raw.get("minimum_components"), "minimum_components", errors, minimum=0)
    required = _boolean_value(raw.get("required"), "required", errors, default=True)
    if not scheme_code: errors.append("scheme_code is required")
    if not component_code: errors.append("component_code is required")
    if not name: errors.append("scheme_name is required")
    if not component_name: errors.append("component_name is required")
    if scheme_version is None: errors.append("scheme_version is required")
    if domain not in {"academic", "coding", "placement"}: errors.append("domain must be academic, coding, or placement")
    if method not in {"weighted_sum", "average", "best_n"}: errors.append("calculation_method is not supported")
    if metric_type not in {"number", "percentage", "integer", "boolean", "short_text", "grade", "rank", "count"}: errors.append("metric_type is not supported")
    if pass_marks is not None and max_marks is not None and pass_marks > max_marks:
        errors.append("pass_marks cannot exceed max_marks")
    existing = db.scalar(select(CollegeAssessmentScheme.id).where(
        CollegeAssessmentScheme.organization_id == run.organization_id,
        CollegeAssessmentScheme.code == scheme_code,
    )) if scheme_code else None
    if existing:
        errors.append("This pattern code already exists; create a new version in Academic Structure instead of implicitly replacing it")
    if scheme_version not in {None, 1}:
        errors.append("A file can create only revision 1; later revisions must clone an existing pattern")
    values = {
        "scheme_code": scheme_code, "scheme_version": scheme_version, "scheme_name": name,
        "domain": domain, "calculation_method": method,
        "final_score_max": final_score_max, "component_code": component_code,
        "component_name": component_name, "metric_type": metric_type,
        "max_marks": max_marks, "weightage_percent": weightage_percent,
        "pass_marks": pass_marks, "required": required,
        "best_n": best_n, "minimum_components": minimum_components,
    }
    return _base_exchange_row(
        run, row_number, action="create", values=values, errors=errors,
        natural_key=f"{scheme_code}:{scheme_version}:{component_code}",
    )


def _stage_cycle_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    scheme_code = str(raw.get("scheme_code") or "").strip().upper()
    scheme_version = _whole_number(raw.get("scheme_version"), "scheme_version", errors, minimum=1)
    cycle_code = _exchange_code(raw.get("cycle_code"), "cycle_code", errors, max_length=60)
    cycle_name = _clean_text(raw.get("cycle_name"), maximum=180)
    scheme = db.scalar(select(CollegeAssessmentScheme).where(
        CollegeAssessmentScheme.organization_id == run.organization_id,
        CollegeAssessmentScheme.code == scheme_code,
        CollegeAssessmentScheme.version_number == scheme_version,
        CollegeAssessmentScheme.status.in_(("active", "frozen")),
    )) if scheme_code and scheme_version else None
    if not scheme: errors.append("An active assessment pattern matching scheme_code and scheme_version was not found")
    component_code = str(raw.get("component_code") or "").strip().upper() or None
    component = db.scalar(select(CollegeAssessmentComponent).where(
        CollegeAssessmentComponent.organization_id == run.organization_id,
        CollegeAssessmentComponent.scheme_id == scheme.id,
        CollegeAssessmentComponent.code == component_code,
    )) if scheme and component_code else None
    if scheme and scheme.domain == "academic" and not component: errors.append("component_code must identify an academic pattern component")
    if component_code and not component: errors.append("component_code does not belong to the selected pattern")
    if not cycle_code: errors.append("cycle_code is required")
    if not cycle_name: errors.append("cycle_name is required")
    if cycle_code and db.scalar(select(CollegeExamCycle.id).where(
        CollegeExamCycle.organization_id == run.organization_id,
        CollegeExamCycle.code == cycle_code,
    )): errors.append("cycle_code already exists")
    def target_ids(value: Any) -> list[str]:
        items = value if isinstance(value, list) else str(value or "").split(",")
        return list(dict.fromkeys(str(item).strip() for item in items if str(item).strip()))

    offering_ids = target_ids(raw.get("offering_ids"))
    cohort_ids = target_ids(raw.get("cohort_ids"))
    if scheme and scheme.domain == "academic" and not offering_ids: errors.append("Academic cycles require at least one offering ID")
    if scheme and scheme.domain != "academic" and not cohort_ids: errors.append("Coding and placement cycles require at least one cohort ID")
    if offering_ids:
        found = set(db.scalars(select(CollegeCourseOffering.id).where(
            CollegeCourseOffering.organization_id == run.organization_id,
            CollegeCourseOffering.id.in_(offering_ids),
        )))
        if found != set(offering_ids): errors.append("One or more offering IDs were not found in this college")
    if cohort_ids:
        found = set(db.scalars(select(CollegeCohort.id).where(
            CollegeCohort.organization_id == run.organization_id,
            CollegeCohort.id.in_(cohort_ids),
        )))
        if found != set(cohort_ids): errors.append("One or more cohort IDs were not found in this college")
    term_id = str(raw.get("term_id") or "").strip() or None
    if term_id and not db.scalar(select(CollegeTerm.id).where(CollegeTerm.id == term_id, CollegeTerm.organization_id == run.organization_id)):
        errors.append("term_id was not found in this college")
    held_on = _date_or_datetime(raw.get("held_on"), "held_on", errors)
    due_on = _date_or_datetime(raw.get("due_on"), "due_on", errors)
    if held_on and due_on and due_on < held_on: errors.append("due_on cannot be before held_on")
    values = {
        "scheme_id": scheme.id if scheme else None, "scheme_component_id": component.id if component else None,
        "term_id": term_id, "cycle_code": cycle_code, "cycle_name": cycle_name,
        "held_on": held_on.isoformat() if held_on else None, "due_on": due_on.isoformat() if due_on else None,
        "offering_ids": offering_ids, "cohort_ids": cohort_ids,
    }
    return _base_exchange_row(run, row_number, action="create", values=values, errors=errors, natural_key=cycle_code or str(row_number))


def _student_by_admission(db: Session, organization_id: str, admission_number: Any):
    value = str(admission_number or "").strip()
    return db.scalar(select(CollegeStudentProfile).where(
        CollegeStudentProfile.organization_id == organization_id,
        CollegeStudentProfile.admission_number == value,
    )) if value else None


def _stage_coding_account_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    student = _student_by_admission(db, run.organization_id, raw.get("admission_number"))
    if not student: errors.append("admission_number was not found")
    allowed = (run.mapping or {}).get("allowed_student_ids")
    if student and allowed is not None and student.id not in set(allowed): errors.append("student is outside your College access")
    platform = str(raw.get("platform") or "").strip().casefold()
    username = str(raw.get("username") or "").strip()
    consent = str(raw.get("consent_status") or "pending").strip().casefold()
    verification = str(raw.get("verification_status") or "unverified").strip().casefold()
    if not platform: errors.append("platform is required")
    if len(username) < 2: errors.append("username is required")
    if consent not in {"pending", "granted", "revoked"}: errors.append("consent_status must be pending, granted, or revoked")
    if verification not in {"unverified", "verified", "failed"}: errors.append("verification_status must be unverified, verified, or failed")
    existing = db.scalar(select(CollegeCodingAccount.id).where(
        CollegeCodingAccount.organization_id == run.organization_id,
        or_(
            and_(CollegeCodingAccount.platform == platform, CollegeCodingAccount.username == username),
            and_(CollegeCodingAccount.student_profile_id == (student.id if student else ""), CollegeCodingAccount.platform == platform),
        ),
    )) if platform and username else None
    if existing: errors.append("This coding account already exists; file uploads never replace it implicitly")
    values = {"student_profile_id": student.id if student else None, "platform": platform, "username": username, "consent_status": consent, "verification_status": verification}
    return _base_exchange_row(run, row_number, action="create", values=values, errors=errors, natural_key=f"{platform}:{username}")


def _stage_company_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    name = _clean_text(raw.get("name"), maximum=200)
    if not name or len(name) < 2: errors.append("Company name is required")
    if name and db.scalar(select(CollegePlacementCompany.id).where(
        CollegePlacementCompany.organization_id == run.organization_id,
        func.lower(CollegePlacementCompany.name) == name.casefold(),
    )): errors.append("This company already exists; uploads do not replace records implicitly")
    values = {field: _clean_text(raw.get(field), maximum=limit) for field, limit in {
        "name": 200, "industry": 100, "website": 500, "contact_name": 160,
        "contact_email": 255, "contact_phone": 40,
    }.items()}
    return _base_exchange_row(run, row_number, action="create", values=values, errors=errors, natural_key=(name or str(row_number)).casefold())


def _stage_drive_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    company_id = str(raw.get("company_id") or "").strip()
    company = db.scalar(select(CollegePlacementCompany).where(
        CollegePlacementCompany.organization_id == run.organization_id,
        CollegePlacementCompany.id == company_id,
        CollegePlacementCompany.is_active.is_(True),
    )) if company_id else None
    if not company: errors.append("company_id was not found or is inactive")
    title = _clean_text(raw.get("title"), maximum=220)
    if not title or len(title) < 2: errors.append("Drive title is required")
    opportunity_type = str(raw.get("opportunity_type") or "campus_drive").strip().casefold()
    status_value = str(raw.get("status") or "draft").strip().casefold()
    if opportunity_type not in {"campus_drive", "internship", "off_campus", "apprenticeship"}: errors.append("opportunity_type is invalid")
    if status_value not in {"draft", "published", "active", "closed", "cancelled"}: errors.append("status is invalid")
    opens_at = _date_or_datetime(raw.get("opens_at"), "opens_at", errors, with_time=True)
    deadline_at = _date_or_datetime(raw.get("deadline_at"), "deadline_at", errors, with_time=True)
    drive_at = _date_or_datetime(raw.get("drive_at"), "drive_at", errors, with_time=True)
    if opens_at and deadline_at and deadline_at < opens_at: errors.append("deadline_at cannot be before opens_at")
    if deadline_at and drive_at and drive_at < deadline_at: errors.append("drive_at cannot be before deadline_at")
    package_min = _whole_number(raw.get("package_min_paise"), "package_min_paise", errors)
    package_max = _whole_number(raw.get("package_max_paise"), "package_max_paise", errors)
    if package_min is not None and package_max is not None and package_max < package_min: errors.append("package_max_paise cannot be below package_min_paise")
    values = {
        "company_id": company_id, "title": title, "opportunity_type": opportunity_type, "status": status_value,
        "opens_at": opens_at.isoformat() if opens_at else None, "deadline_at": deadline_at.isoformat() if deadline_at else None,
        "drive_at": drive_at.isoformat() if drive_at else None, "work_location": _clean_text(raw.get("work_location"), maximum=180),
        "employment_type": _clean_text(raw.get("employment_type"), maximum=50),
        "package_min_paise": package_min, "package_max_paise": package_max,
    }
    return _base_exchange_row(run, row_number, action="create", values=values, errors=errors, natural_key=f"{company_id}:{title or row_number}")


def _stage_application_row(db: Session, run: DataExchangeRun, row_number: int, raw: dict) -> DataExchangeRow:
    errors: list[str] = []
    record_id = str(raw.get("record_id") or "").strip()
    application = db.scalar(select(CollegePlacementApplication).where(
        CollegePlacementApplication.id == record_id,
        CollegePlacementApplication.organization_id == run.organization_id,
    )) if record_id else None
    if not application: errors.append("Application ID was not found")
    version = _whole_number(raw.get("version"), "version", errors, minimum=1)
    if application and version != application.version: errors.append("Application version is stale; download a new update template")
    allowed = (run.mapping or {}).get("allowed_student_ids")
    if application and allowed is not None and application.student_profile_id not in set(allowed): errors.append("Application is outside your College access")
    stage_code = str(raw.get("stage_code") or "").strip().casefold()
    stage = db.scalar(select(CollegePipelineStage).where(
        CollegePipelineStage.organization_id == run.organization_id,
        CollegePipelineStage.slug == stage_code,
        CollegePipelineStage.is_enabled.is_(True),
    )) if stage_code else None
    if not stage: errors.append("stage_code was not found or is disabled")
    reason = _clean_text(raw.get("reason"), maximum=2000)
    if not reason or len(reason) < 2: errors.append("A stage-change reason is required")
    if application and stage and application.current_stage_id == stage.id: errors.append("Application is already in this stage")
    values = {"stage_id": stage.id if stage else None, "stage_code": stage_code, "reason": reason}
    return _base_exchange_row(
        run, row_number, action="update", values=values, errors=errors,
        natural_key=record_id or str(row_number), record_id=record_id or None,
        record_version=version,
        current_values={"stage_id": application.current_stage_id if application else None, "version": application.version if application else None},
        changes={"stage_id": stage.id if stage else None},
    )


def stage_exchange_rows(db: Session, run: DataExchangeRun, rows: list[dict]) -> None:
    exchange_rows: list[DataExchangeRow] = []
    legacy_rows: list[dict] = []
    for index, raw in enumerate(rows, start=1):
        marker = raw.get("_row_number")
        row_number = index
        if marker not in (None, ""):
            marker_errors: list[str] = []
            parsed_marker = _whole_number(marker, "row number", marker_errors, minimum=1)
            if parsed_marker is not None and not marker_errors:
                row_number = parsed_marker
        key = run.resource_key
        if key == "academic_structure":
            key = str(raw.get("_resource_key") or "").casefold()
            aliases = {
                "graduation_batches_and_sections": "cohorts",
                "academic_years_and_terms": "terms",
                "course_offerings": "offerings",
                "assessment_patterns": "assessment_schemes",
            }
            key = aliases.get(key, key)
        if key in STRUCTURE_MODELS:
            exchange_rows.append(_stage_structure_row(db, run, row_number, raw, key))
        elif key == "assessment_marks":
            exchange_rows.append(_stage_marks_row(db, run, row_number, raw))
        elif key == "assessment_schemes":
            staged = _stage_scheme_row(db, run, row_number, raw)
            if run.resource_key == "academic_structure":
                staged.values = {**(staged.values or {}), "_resource_key": "assessment_schemes"}
            exchange_rows.append(staged)
        elif key == "exam_cycles":
            exchange_rows.append(_stage_cycle_row(db, run, row_number, raw))
        elif key == "coding_accounts":
            exchange_rows.append(_stage_coding_account_row(db, run, row_number, raw))
        elif key == "companies":
            exchange_rows.append(_stage_company_row(db, run, row_number, raw))
        elif key == "drives":
            exchange_rows.append(_stage_drive_row(db, run, row_number, raw))
        elif key == "applications":
            exchange_rows.append(_stage_application_row(db, run, row_number, raw))
        elif key in RESOURCE_FIELDS:
            legacy_rows.append(raw)
        else:
            exchange_rows.append(DataExchangeRow(
                organization_id=run.organization_id, run_id=run.id, row_number=row_number,
                status="invalid", errors=["This resource currently supports generated templates and schemas, but not file commit"],
                values=raw,
            ))
    if legacy_rows:
        legacy = stage_rows(
            db,
            organization_id=run.organization_id,
            user_id=run.initiated_by_user_id,
            source_type=run.source_type,
            resource_type=run.resource_key,
            rows=legacy_rows,
            mapping={},
            allowed_student_ids=(set(run.mapping.get("allowed_student_ids")) if (run.mapping or {}).get("allowed_student_ids") is not None else None),
            allowed_program_ids=(set(run.mapping.get("allowed_program_ids")) if (run.mapping or {}).get("allowed_program_ids") is not None else None),
            allowed_cohort_ids=(set(run.mapping.get("allowed_cohort_ids")) if (run.mapping or {}).get("allowed_cohort_ids") is not None else None),
        )
        run.legacy_import_run_id = legacy.id
        for staged in legacy.staged_rows:
            values = staged.get("data") or {}
            action = str(values.get("_manual_action") or "create")
            exchange_rows.append(DataExchangeRow(
                organization_id=run.organization_id,
                run_id=run.id,
                row_number=int(staged["row"]),
                action=action,
                status="valid" if staged.get("valid") else "invalid",
                values=values,
                record_id=values.get("_record_id"),
                record_version=values.get("_record_version"),
                errors=staged.get("errors") or [],
            ))
    scheme_rows = [
        item for item in exchange_rows
        if run.resource_key == "assessment_schemes"
        or (item.values or {}).get("_resource_key") == "assessment_schemes"
    ]
    if scheme_rows:
        groups: dict[str, list[DataExchangeRow]] = {}
        for item in scheme_rows:
            groups.setdefault(str(item.values.get("scheme_code") or ""), []).append(item)
        for group in groups.values():
            component_codes = [str(item.values.get("component_code") or "") for item in group]
            if len(component_codes) != len(set(component_codes)):
                for item in group:
                    item.errors = [*(item.errors or []), "Component codes must be unique within a pattern"]
                    item.status = "invalid"
            signatures = {
                (
                    item.values.get("scheme_name"), item.values.get("domain"),
                    item.values.get("calculation_method"), str(item.values.get("final_score_max")),
                    str(item.values.get("best_n")), str(item.values.get("minimum_components")),
                ) for item in group
            }
            if len(signatures) > 1:
                for item in group:
                    item.errors = [*(item.errors or []), "Every component row for a pattern must repeat the same pattern settings"]
                    item.status = "invalid"
            if all(item.status == "valid" for item in group):
                first = group[0].values
                config = {}
                if first.get("best_n") is not None:
                    config["best_n"] = first["best_n"]
                if first.get("minimum_components") is not None:
                    config["minimum_components"] = first["minimum_components"]
                definitions = [{
                    "name": item.values["component_name"],
                    "code": item.values["component_code"],
                    "component_type": "assessment",
                    "metric_type": item.values["metric_type"],
                    "display_order": order,
                    "max_marks": item.values.get("max_marks"),
                    "weightage_bps": int(Decimal(str(item.values.get("weightage_percent") or 0)) * 100),
                    "pass_marks": item.values.get("pass_marks"),
                    "is_required": _truthy(item.values.get("required"), True),
                    "settings": {},
                } for order, item in enumerate(group, start=1)]
                try:
                    validate_component_definitions(definitions, first["calculation_method"], config)
                except HTTPException as exc:
                    message = str(exc.detail)
                    for item in group:
                        item.errors = [*(item.errors or []), message]
                        item.status = "invalid"
    exchange_rows.sort(key=lambda item: item.row_number)
    db.add_all(exchange_rows)
    db.flush()
    _summarize_exchange_rows(run, exchange_rows)


def _summarize_exchange_rows(run: DataExchangeRun, rows: list[DataExchangeRow]) -> None:
    run.row_count = len(rows)
    run.valid_count = sum(item.status == "valid" for item in rows)
    run.create_count = sum(item.status == "valid" and item.action == "create" for item in rows)
    run.update_count = sum(item.status == "valid" and item.action == "update" for item in rows)
    run.unchanged_count = sum(item.status == "valid" and item.action == "unchanged" for item in rows)
    run.invalid_count = sum(item.status == "invalid" for item in rows)
    run.conflict_count = sum(item.status == "conflict" for item in rows)
    run.status = "ready" if run.valid_count else "invalid"


def _parent_by_code(db: Session, model, organization_id: str, code: str):
    row = db.scalar(select(model).where(
        model.organization_id == organization_id,
        func.upper(model.code) == str(code).strip().upper().replace(" ", "-"),
    ))
    if not row:
        raise ValueError(f"Parent code {code} was not found")
    return row


def _apply_structure_row(db: Session, run: DataExchangeRun, item: DataExchangeRow) -> None:
    key = str(item.values.get("_resource_key") or run.resource_key)
    values = {k: v for k, v in item.values.items() if not k.startswith("_") and k not in {"external_id", "source_updated_at"}}
    model = STRUCTURE_MODELS[key]
    record = db.get(model, item.record_id) if item.record_id else None
    if record and (record.organization_id != run.organization_id or record.version != item.record_version):
        raise ValueError(f"Row {item.row_number} changed after preview")
    if key == "programs":
        values["department_id"] = _parent_by_code(db, CollegeDepartment, run.organization_id, values.pop("department_code")).id
    elif key == "cohorts":
        values["program_id"] = _parent_by_code(db, CollegeProgram, run.organization_id, values.pop("program_code")).id
    elif key == "courses":
        values["department_id"] = _parent_by_code(db, CollegeDepartment, run.organization_id, values.pop("department_code")).id
    elif key == "terms":
        values["starts_on"] = date.fromisoformat(str(values["starts_on"])[:10])
        values["ends_on"] = date.fromisoformat(str(values["ends_on"])[:10])
    if record:
        for field_name, value in values.items():
            if hasattr(record, field_name):
                setattr(record, field_name, value)
        record.version += 1
    else:
        record = model(organization_id=run.organization_id, **values)
        db.add(record)
    db.flush()
    item.record_id = record.id
    item.record_version = getattr(record, "version", None)
    item.status = "committed"


def _apply_marks_row(db: Session, run: DataExchangeRun, item: DataExchangeRow, user_id: str | None, can_correct: bool) -> None:
    values = item.values
    assessment = db.get(CollegeAssessment, values["assessment_id"])
    if not assessment or assessment.organization_id != run.organization_id:
        raise ValueError(f"Row {item.row_number} assessment is no longer available")
    score = db.get(CollegeAssessmentScore, item.record_id) if item.record_id else None
    if score and (score.organization_id != run.organization_id or score.version != item.record_version):
        raise ValueError(f"Row {item.row_number} changed after preview")
    if assessment.status == "published":
        if not can_correct:
            raise PermissionError("Published results require the assessment correction permission")
        if not run.correction_reason or len(run.correction_reason.strip()) < 3:
            raise ValueError("A correction reason is required for published results")
    metrics = dict(values.get("metrics") or {})
    cycle = db.get(CollegeExamCycle, assessment.exam_cycle_id) if assessment.exam_cycle_id else None
    calculated = None
    first_value = next(iter(metrics.values()), None) if len(assessment.metric_schema or []) == 1 else None
    before = {"metrics": dict(score.metrics or {}), "calculated_score": float(score.calculated_score) if score and score.calculated_score is not None else None} if score else None
    if score:
        score.metrics = metrics
        score.calculated_score = calculated
        score.marks_awarded = first_value if isinstance(first_value, (int, float)) else score.marks_awarded
        score.graded_by_user_id = user_id
        score.version += 1
    else:
        score = CollegeAssessmentScore(
            organization_id=run.organization_id,
            assessment_id=assessment.id,
            student_profile_id=values["student_profile_id"],
            metrics=metrics,
            calculated_score=calculated,
            marks_awarded=first_value if isinstance(first_value, (int, float)) else None,
            graded_by_user_id=user_id,
        )
        db.add(score)
    db.flush()
    if cycle:
        calculated = recalculate_assessment_score(db, assessment, values["student_profile_id"])
        score.calculated_score = calculated
    item.record_id = score.id
    item.record_version = score.version
    item.current_values = {"before": before, "after": {"metrics": metrics, "calculated_score": float(calculated) if calculated is not None else None}}
    item.status = "committed"


def _truthy(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().casefold() in {"true", "yes", "1", "required"}


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{value} is not a valid number") from exc
    if not result.is_finite():
        raise ValueError("Numeric values must be finite")
    return result


def _apply_scheme_rows(db: Session, run: DataExchangeRun, rows: list[DataExchangeRow]) -> None:
    groups: dict[str, list[DataExchangeRow]] = {}
    for item in rows:
        groups.setdefault(str(item.values["scheme_code"]), []).append(item)
    for scheme_code, group in groups.items():
        first = group[0].values
        components = []
        for order, item in enumerate(group, start=1):
            values = item.values
            weight = _decimal_or_none(values.get("weightage_percent")) or Decimal("0")
            components.append({
                "name": values["component_name"], "code": values["component_code"],
                "component_type": "assessment", "metric_type": values["metric_type"],
                "display_order": order, "max_marks": _decimal_or_none(values.get("max_marks")),
                "weightage_bps": int(weight * 100), "pass_marks": _decimal_or_none(values.get("pass_marks")),
                "is_required": _truthy(values.get("required"), True), "aggregation_group": None,
                "settings": {},
            })
        config = {}
        if first["calculation_method"] == "best_n":
            best_n = int(first.get("best_n") or 0)
            if not best_n:
                raise ValueError(f"Pattern {scheme_code} requires best_n")
            config["best_n"] = best_n
        minimum_components = int(first.get("minimum_components") or 0)
        if minimum_components:
            config["minimum_components"] = minimum_components
        definitions = validate_component_definitions(components, first["calculation_method"], config)
        final_score_max = _decimal_or_none(first.get("final_score_max"))
        if final_score_max is None or final_score_max <= 0:
            raise ValueError(f"Pattern {scheme_code} needs a positive final score scale")
        scheme = CollegeAssessmentScheme(
            organization_id=run.organization_id, name=first["scheme_name"], code=scheme_code,
            domain=first["domain"], version_number=1, status="draft",
            final_score_max=final_score_max, calculation_method=first["calculation_method"],
            calculation_config=config,
        )
        db.add(scheme)
        db.flush()
        component_rows = [CollegeAssessmentComponent(
            organization_id=run.organization_id, scheme_id=scheme.id, **definition,
        ) for definition in definitions]
        db.add_all(component_rows)
        db.flush()
        for item in group:
            item.record_id = scheme.id
            item.record_version = scheme.version
            item.status = "committed"
            run.committed_count += 1


def _apply_cycle_row(db: Session, run: DataExchangeRun, item: DataExchangeRow) -> None:
    values = item.values
    scheme = db.scalar(select(CollegeAssessmentScheme).where(
        CollegeAssessmentScheme.id == values["scheme_id"],
        CollegeAssessmentScheme.organization_id == run.organization_id,
    ))
    if not scheme:
        raise ValueError(f"Row {item.row_number} assessment pattern is no longer available")
    components = list(db.scalars(select(CollegeAssessmentComponent).where(
        CollegeAssessmentComponent.organization_id == run.organization_id,
        CollegeAssessmentComponent.scheme_id == scheme.id,
    ).order_by(CollegeAssessmentComponent.display_order)))
    component = next((row for row in components if row.id == values.get("scheme_component_id")), None)
    snapshot = build_scheme_snapshot(scheme, components)
    cycle = CollegeExamCycle(
        organization_id=run.organization_id, scheme_id=scheme.id,
        scheme_component_id=component.id if component else None, term_id=values.get("term_id"),
        name=values["cycle_name"], code=values["cycle_code"], domain=scheme.domain,
        held_on=date.fromisoformat(values["held_on"]) if values.get("held_on") else None,
        due_on=date.fromisoformat(values["due_on"]) if values.get("due_on") else None,
        target_offering_ids=values.get("offering_ids") or [],
        target_cohort_ids=values.get("cohort_ids") or [], scheme_snapshot=snapshot,
    )
    db.add(cycle)
    db.flush()
    metric_schema = [component_payload(component)] if component else snapshot["components"]
    assessments = []
    if scheme.domain == "academic":
        offerings = list(db.scalars(select(CollegeCourseOffering).where(
            CollegeCourseOffering.organization_id == run.organization_id,
            CollegeCourseOffering.id.in_(values.get("offering_ids") or []),
        )))
        for offering in offerings:
            assessments.append(CollegeAssessment(
                organization_id=run.organization_id, offering_id=offering.id,
                cohort_id=offering.cohort_id, exam_cycle_id=cycle.id, scheme_id=scheme.id,
                scheme_component_id=component.id, title=cycle.name,
                assessment_type=component.component_type,
                max_marks=component.max_marks or scheme.final_score_max,
                weightage_bps=component.weightage_bps, due_on=cycle.due_on,
                metric_schema=metric_schema,
            ))
    else:
        for cohort_id in values.get("cohort_ids") or []:
            assessments.append(CollegeAssessment(
                organization_id=run.organization_id, cohort_id=cohort_id,
                exam_cycle_id=cycle.id, scheme_id=scheme.id,
                scheme_component_id=component.id if component else None,
                title=cycle.name, assessment_type=component.component_type if component else scheme.domain,
                max_marks=component.max_marks if component and component.max_marks else scheme.final_score_max,
                weightage_bps=component.weightage_bps if component else 10000,
                due_on=cycle.due_on, metric_schema=metric_schema,
            ))
    db.add_all(assessments)
    freeze_scheme(scheme)
    db.flush()
    item.record_id = cycle.id
    item.record_version = cycle.version
    item.status = "committed"


def _apply_simple_exchange_row(db: Session, run: DataExchangeRun, item: DataExchangeRow, user_id: str | None) -> None:
    values = dict(item.values)
    if run.resource_key == "coding_accounts":
        record = CollegeCodingAccount(organization_id=run.organization_id, **values)
    elif run.resource_key == "companies":
        record = CollegePlacementCompany(organization_id=run.organization_id, **values)
    elif run.resource_key == "drives":
        for field_name in ("opens_at", "deadline_at", "drive_at"):
            values[field_name] = datetime.fromisoformat(values[field_name]) if values.get(field_name) else None
        record = CollegePlacementOpportunity(
            organization_id=run.organization_id, eligibility_rules={}, rounds=[],
            owner_user_id=user_id, **values,
        )
    elif run.resource_key == "applications":
        record = db.get(CollegePlacementApplication, item.record_id)
        if not record or record.organization_id != run.organization_id or record.version != item.record_version:
            raise ValueError(f"Row {item.row_number} application changed after preview")
        stage = db.get(CollegePipelineStage, values["stage_id"])
        if not stage or stage.organization_id != run.organization_id or not stage.is_enabled:
            raise ValueError(f"Row {item.row_number} stage is no longer available")
        previous = record.current_stage_id
        record.current_stage_id = stage.id
        record.version += 1
        if stage.slug == "applied" and not record.applied_at:
            record.applied_at = datetime.now(timezone.utc)
        if stage.stage_type != "active":
            record.outcome = stage.stage_type
        db.add(CollegeApplicationStageEvent(
            organization_id=run.organization_id, application_id=record.id,
            from_stage_id=previous, to_stage_id=stage.id,
            changed_by_user_id=user_id, reason=values["reason"],
            occurred_at=datetime.now(timezone.utc),
        ))
    else:
        raise ValueError(f"{run.resource_key} does not have a file committer")
    db.add(record)
    db.flush()
    item.record_id = record.id
    item.record_version = getattr(record, "version", None)
    item.status = "committed"


def commit_exchange_run(db: Session, run: DataExchangeRun, *, user_id: str | None, can_correct: bool) -> DataExchangeRun:
    if run.status == "committed":
        return run
    if run.status != "ready":
        raise ValueError("Only a validated run can be committed")
    rows = list(db.scalars(select(DataExchangeRow).where(
        DataExchangeRow.run_id == run.id,
        DataExchangeRow.organization_id == run.organization_id,
        DataExchangeRow.status == "valid",
    ).order_by(DataExchangeRow.row_number)))
    if run.resource_key == "academic_structure":
        order = {key: index for index, key in enumerate(STRUCTURE_PACKAGE_ORDER)}
        rows.sort(key=lambda item: (order.get(str(item.values.get("_resource_key")), 999), item.row_number))
    if run.legacy_import_run_id:
        legacy = db.get(CollegeImportRun, run.legacy_import_run_id)
        if legacy:
            commit_run(db, legacy)
            run.committed_count += legacy.committed_count
            for item in rows:
                item.status = "committed"
    elif run.resource_key == "assessment_schemes":
        _apply_scheme_rows(db, run, rows)
    else:
        scheme_rows = [
            item for item in rows
            if run.resource_key == "academic_structure"
            and (item.values or {}).get("_resource_key") == "assessment_schemes"
        ]
        regular_rows = [item for item in rows if item not in scheme_rows]
        for item in regular_rows:
            if item.action == "unchanged":
                item.status = "committed"
                continue
            if run.resource_key == "assessment_marks":
                _apply_marks_row(db, run, item, user_id, can_correct)
            elif run.resource_key == "exam_cycles":
                _apply_cycle_row(db, run, item)
            elif run.resource_key in {"coding_accounts", "companies", "drives", "applications"}:
                _apply_simple_exchange_row(db, run, item, user_id)
            else:
                _apply_structure_row(db, run, item)
            run.committed_count += 1
        if scheme_rows:
            _apply_scheme_rows(db, run, scheme_rows)
    run.status = "committed"
    run.completed_at = datetime.now(timezone.utc)
    return run


def ingest_exchange_records(
    db: Session,
    *,
    organization_id: str,
    resource_key: str,
    records: list[dict],
    source_type: str,
    idempotency_key: str,
    request_hash: str,
    initiated_by_user_id: str | None = None,
    access_mapping: dict | None = None,
    auto_commit: bool = True,
) -> DataExchangeRun:
    """Stage structured ERP/API records through the same registry as files."""
    if resource_key not in RESOURCES or resource_key in {"academic_structure", "assessment_marks"}:
        raise ValueError("This resource needs its dedicated exchange handler")
    existing = db.scalar(select(DataExchangeRun).where(
        DataExchangeRun.organization_id == organization_id,
        DataExchangeRun.idempotency_key == idempotency_key,
    ))
    if existing:
        if existing.request_hash != request_hash or existing.resource_key != resource_key:
            raise ValueError("The idempotency key was already used with different data")
        return existing
    run = DataExchangeRun(
        organization_id=organization_id,
        resource_key=resource_key,
        operation="import",
        source_type=source_type,
        status="uploaded",
        schema_version="1",
        scope={},
        mapping=access_mapping or {},
        idempotency_key=idempotency_key,
        request_hash=request_hash,
        initiated_by_user_id=initiated_by_user_id,
        started_at=datetime.now(timezone.utc),
    )
    db.add(run)
    db.flush()
    stage_exchange_rows(db, run, records)
    if auto_commit and run.valid_count:
        commit_exchange_run(db, run, user_id=initiated_by_user_id, can_correct=False)
    return run


def ingest_assessment_metric_records(
    db: Session,
    *,
    organization_id: str,
    records: list[dict],
    source_type: str,
    idempotency_key: str,
    request_hash: str,
    initiated_by_user_id: str | None = None,
    access_mapping: dict | None = None,
    auto_commit: bool = True,
) -> DataExchangeRun:
    """Stage dynamic ERP/API marks against one immutable exam-cycle snapshot."""
    existing_run = db.scalar(select(DataExchangeRun).where(
        DataExchangeRun.organization_id == organization_id,
        DataExchangeRun.idempotency_key == idempotency_key,
    ))
    if existing_run:
        if existing_run.request_hash != request_hash:
            raise ValueError("The idempotency key was already used with different assessment data")
        return existing_run
    cycle_codes = {
        _exchange_code(row.get("cycle_code"), "cycle_code", [], max_length=60)
        for row in records
        if row.get("cycle_code") not in (None, "")
    }
    if len(cycle_codes) != 1 or len(records) == 0 or any(row.get("cycle_code") in (None, "") for row in records):
        raise ValueError("An assessment-marks batch must identify exactly one cycle_code")
    cycle_code = next(iter(cycle_codes))
    cycle = db.scalar(select(CollegeExamCycle).where(
        CollegeExamCycle.organization_id == organization_id,
        CollegeExamCycle.code == cycle_code,
    ))
    if not cycle:
        raise ValueError("cycle_code was not found")

    run = DataExchangeRun(
        organization_id=organization_id,
        resource_key="assessment_marks",
        operation="import",
        source_type=source_type,
        status="uploaded",
        schema_version=str((cycle.scheme_snapshot or {}).get("scheme_version") or "1"),
        scope={"cycle_id": cycle.id},
        mapping=access_mapping or {},
        idempotency_key=idempotency_key,
        request_hash=request_hash,
        initiated_by_user_id=initiated_by_user_id,
        started_at=datetime.now(timezone.utc),
    )
    db.add(run)
    db.flush()

    snapshot = cycle.scheme_snapshot or {}
    expected_scheme_code = str(snapshot.get("scheme_code") or "").upper()
    expected_scheme_version = int(snapshot.get("scheme_version") or 0)
    transformed: list[dict] = []
    invalid_rows: list[DataExchangeRow] = []
    seen_keys: set[str] = set()
    for index, raw in enumerate(records, start=1):
        errors: list[str] = []
        scheme_code = str(raw.get("scheme_code") or "").strip().upper()
        scheme_version = _whole_number(raw.get("scheme_version"), "scheme_version", errors, minimum=1)
        if scheme_code != expected_scheme_code or scheme_version != expected_scheme_version:
            errors.append("scheme_code and scheme_version must match the selected cycle snapshot")

        student_value = raw.get("student")
        admission_number = str(
            student_value.get("admission_number") if isinstance(student_value, dict)
            else student_value or raw.get("admission_number") or ""
        ).strip()
        student = _student_by_admission(db, organization_id, admission_number)
        if not student:
            errors.append("student admission number was not found")
        allowed_student_ids = (access_mapping or {}).get("allowed_student_ids")
        if student and allowed_student_ids is not None and student.id not in set(allowed_student_ids):
            errors.append("student is outside the authorized College scope")

        academic_scope = raw.get("academic_scope") if isinstance(raw.get("academic_scope"), dict) else {}
        assessment_id = str(raw.get("assessment_id") or academic_scope.get("assessment_id") or "").strip()
        offering_id = str(raw.get("offering_id") or academic_scope.get("offering_id") or "").strip()
        candidates = []
        if student:
            statement = select(CollegeAssessment).where(
                CollegeAssessment.organization_id == organization_id,
                CollegeAssessment.exam_cycle_id == cycle.id,
                CollegeAssessment.cohort_id == student.cohort_id,
            )
            if assessment_id:
                statement = statement.where(CollegeAssessment.id == assessment_id)
            if offering_id:
                statement = statement.where(CollegeAssessment.offering_id == offering_id)
            candidates = list(db.scalars(statement))
        assessment = candidates[0] if len(candidates) == 1 else None
        if not assessment:
            if len(candidates) > 1:
                errors.append("academic_scope must identify assessment_id or offering_id for this student")
            else:
                errors.append("the cycle does not contain an assessment for this student and academic scope")

        metrics = raw.get("metrics")
        if not isinstance(metrics, dict):
            errors.append("metrics must be an object")
            metrics = {}
        natural_key = f"{assessment.id if assessment else assessment_id}:{admission_number}"
        if assessment and natural_key in seen_keys:
            errors.append("the batch contains this student and assessment more than once")
        seen_keys.add(natural_key)

        existing = db.scalar(select(CollegeAssessmentScore).where(
            CollegeAssessmentScore.organization_id == organization_id,
            CollegeAssessmentScore.assessment_id == assessment.id,
            CollegeAssessmentScore.student_profile_id == student.id,
        )) if assessment and student else None
        if existing and assessment.status == "published":
            errors.append("published marks require an authorized reviewed correction")

        if errors:
            invalid_rows.append(DataExchangeRow(
                organization_id=organization_id,
                run_id=run.id,
                row_number=index,
                action="update" if existing else "create",
                status="invalid",
                natural_key=natural_key,
                record_id=existing.id if existing else None,
                record_version=existing.version if existing else None,
                values={
                    "cycle_code": cycle_code,
                    "scheme_code": scheme_code,
                    "scheme_version": scheme_version,
                    "admission_number": admission_number,
                    "assessment_id": assessment.id if assessment else assessment_id or None,
                    "metrics": metrics,
                },
                errors=list(dict.fromkeys(errors)),
            ))
            continue
        transformed.append({
            "_row_number": index,
            "assessment_id": assessment.id,
            "record_id": existing.id if existing else None,
            "version": existing.version if existing else None,
            "admission_number": admission_number,
            **metrics,
        })

    stage_exchange_rows(db, run, transformed)
    if invalid_rows:
        db.add_all(invalid_rows)
        db.flush()
    all_rows = list(db.scalars(select(DataExchangeRow).where(
        DataExchangeRow.organization_id == organization_id,
        DataExchangeRow.run_id == run.id,
    ).order_by(DataExchangeRow.row_number)))
    _summarize_exchange_rows(run, all_rows)
    if auto_commit and run.valid_count:
        commit_exchange_run(db, run, user_id=initiated_by_user_id, can_correct=False)
    return run


def correction_workbook(run: DataExchangeRun, rows: Iterable[DataExchangeRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Corrections"
    fields: list[str] = []
    row_list = list(rows)
    for item in row_list:
        for key in item.values:
            if not key.startswith("_") and key not in fields:
                fields.append(key)
    columns = ["row_number", "errors", *fields]
    sheet.append(columns)
    for item in row_list:
        sheet.append([item.row_number, "; ".join(item.errors or []), *[item.values.get(key) for key in fields]])
    _format_sheet(sheet, columns)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def process_import_job(db: Session, run_id: str) -> None:
    run = db.get(DataExchangeRun, run_id)
    if not run or run.status not in {"queued", "validating"}:
        return
    source = db.scalar(select(DataExchangeArtifact).where(
        DataExchangeArtifact.run_id == run.id,
        DataExchangeArtifact.kind == "source",
    ))
    if not source:
        raise ValueError("The uploaded source artifact is unavailable")
    run.status = "validating"
    rows, metadata = parse_upload(source.content, source.filename)
    if metadata.get("resource_key") and metadata["resource_key"] != run.resource_key:
        raise ValueError("The workbook resource does not match this import")
    stage_exchange_rows(db, run, rows)
    invalid = list(db.scalars(select(DataExchangeRow).where(
        DataExchangeRow.run_id == run.id,
        DataExchangeRow.status.in_(("invalid", "conflict")),
    ).order_by(DataExchangeRow.row_number)))
    if invalid:
        content = correction_workbook(run, invalid)
        store_artifact(
            db, run, kind="corrections", filename=f"{run.resource_key}-corrections.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=content,
        )


def process_export_job(db: Session, run_id: str) -> None:
    run = db.get(DataExchangeRun, run_id)
    if not run or run.status not in {"queued", "exporting"}:
        return
    run.status = "exporting"
    content, filename, content_type, row_count = generate_export(
        db, run.organization_id, run.resource_key,
        file_format=run.file_format or "xlsx", scope=run.scope or {},
    )
    store_artifact(db, run, kind="export", filename=filename, content_type=content_type, content=content)
    run.row_count = row_count
    run.status = "completed"
    run.completed_at = datetime.now(timezone.utc)
