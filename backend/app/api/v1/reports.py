"""Permission-scoped business reports and exports."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import false, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import require_entitlements, require_permissions
from app.models import SaleInvoice, SalePayment
from app.services.business_access import (
    allowed_client_ids,
    ensure_location,
    filter_locations,
    organization_for,
)
from app.services.entitlements import entitlement_value
from app.services.rbac import get_user_permissions

router = APIRouter(prefix="/reports", tags=["reports"])


def _period_bounds(org_timezone: str, start: date, end: date) -> tuple[datetime, datetime]:
    if start > end:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The start date must be before the end date")
    if (end - start).days > 731:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Choose a date range of two years or less")
    tz = ZoneInfo(org_timezone)
    start_at = datetime.combine(start, time.min, tzinfo=tz).astimezone(timezone.utc)
    end_at = datetime.combine(end + timedelta(days=1), time.min, tzinfo=tz).astimezone(timezone.utc)
    return start_at, end_at


def _scope_invoices(statement, db: Session, user, location_id: str | None):
    statement = filter_locations(statement, SaleInvoice, db, user)
    if location_id:
        ensure_location(db, user, location_id)
        statement = statement.where(SaleInvoice.location_id == location_id)
    clients = allowed_client_ids(db, user)
    if clients is not None:
        statement = statement.where(SaleInvoice.client_id.in_(clients) if clients else false())
    return statement


def sales_rows(db: Session, user, start_at: datetime, end_at: datetime, location_id: str | None):
    statement = select(SaleInvoice).where(
        SaleInvoice.organization_id == user.organization_id,
        SaleInvoice.created_at >= start_at,
        SaleInvoice.created_at < end_at,
    )
    statement = _scope_invoices(statement, db, user, location_id)
    return db.execute(statement.order_by(SaleInvoice.created_at, SaleInvoice.id)).scalars().all()


def payment_rows(db: Session, user, start_at: datetime, end_at: datetime, location_id: str | None):
    statement = (
        select(SalePayment.created_at, SalePayment.amount_paise)
        .join(SaleInvoice, SaleInvoice.id == SalePayment.invoice_id)
        .where(
            SalePayment.organization_id == user.organization_id,
            SalePayment.status == "captured",
            SalePayment.created_at >= start_at,
            SalePayment.created_at < end_at,
        )
    )
    statement = _scope_invoices(statement, db, user, location_id)
    return db.execute(statement.order_by(SalePayment.created_at)).all()


def _change(current: int, previous: int) -> float | None:
    if previous == 0:
        return None
    return round((current - previous) * 100 / previous, 1)


def _metric(identifier: str, label: str, value: int, previous: int, *, format: str = "number", tone: str = "neutral") -> dict:
    return {
        "id": identifier,
        "label": label,
        "value": value,
        "format": format,
        "tone": tone,
        "comparison": {"previous": previous, "change_percent": _change(value, previous)},
        "destination": {"route": "sales"},
    }


def _summary_values(invoices, payments) -> dict:
    billed = sum(int(row.total_paise) for row in invoices)
    collected = sum(int(row.amount_paise) for row in payments)
    outstanding = sum(max(int(row.total_paise) - int(row.paid_paise), 0) for row in invoices)
    return {
        "invoice_count": len(invoices),
        "billed_paise": billed,
        "collected_paise": collected,
        "outstanding_paise": outstanding,
    }


def _daily_series(invoices, payments, start: date, end: date, org_timezone: str) -> list[dict]:
    tz = ZoneInfo(org_timezone)
    billed_by_day: dict[date, int] = defaultdict(int)
    collected_by_day: dict[date, int] = defaultdict(int)
    for invoice in invoices:
        billed_by_day[invoice.created_at.astimezone(tz).date()] += int(invoice.total_paise)
    for occurred_at, amount_paise in payments:
        collected_by_day[occurred_at.astimezone(tz).date()] += int(amount_paise)

    points = []
    cursor = start
    while cursor <= end:
        points.append({
            "date": cursor.isoformat(),
            "billed_paise": billed_by_day[cursor],
            "collected_paise": collected_by_day[cursor],
        })
        cursor += timedelta(days=1)
    return points


@router.get("/summary")
def summary(
    start: date | None = None,
    end: date | None = None,
    location_id: str | None = None,
    user=Depends(require_permissions("reports.view")),
    db: Session = Depends(get_db),
):
    org = organization_for(db, user)
    local_today = datetime.now(ZoneInfo(org.timezone)).date()
    start = start or local_today.replace(day=1)
    end = end or local_today
    start_at, end_at = _period_bounds(org.timezone, start, end)

    period_days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=period_days - 1)
    previous_start_at, previous_end_at = _period_bounds(org.timezone, previous_start, previous_end)

    invoices = sales_rows(db, user, start_at, end_at, location_id)
    payments = payment_rows(db, user, start_at, end_at, location_id)
    previous_invoices = sales_rows(db, user, previous_start_at, previous_end_at, location_id)
    previous_payments = payment_rows(db, user, previous_start_at, previous_end_at, location_id)
    values = _summary_values(invoices, payments)
    previous = _summary_values(previous_invoices, previous_payments)

    status_counts: dict[str, int] = defaultdict(int)
    for invoice in invoices:
        status_counts[invoice.status] += 1
    breakdown_items = [
        {"id": invoice_status, "label": invoice_status.replace("_", " ").title(), "value": count}
        for invoice_status, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0]))
    ]
    outstanding_rows = sorted(
        (row for row in invoices if row.total_paise > row.paid_paise and row.status not in {"void", "refunded"}),
        key=lambda row: (row.total_paise - row.paid_paise, row.created_at),
        reverse=True,
    )[:5]
    permissions = get_user_permissions(db, user)
    can_export = "reports.exports" in permissions and bool(entitlement_value(db, org, "reports.exports", False))
    generated_at = datetime.now(timezone.utc)
    is_college = org.industry.value == "college"

    return {
        # Existing fields remain available while consumers move to WorkspaceOverviewV2.
        "start": start,
        "end": end,
        **values,
        "schema_version": 2,
        "source": {
            "generated_at": generated_at,
            "timezone": org.timezone,
            "freshness": "live",
        },
        "metrics": [
            _metric("billed", "Fees invoiced" if is_college else "Billed", values["billed_paise"], previous["billed_paise"], format="money"),
            _metric("collected", "Fees collected" if is_college else "Collected", values["collected_paise"], previous["collected_paise"], format="money"),
            _metric(
                "outstanding", "Outstanding", values["outstanding_paise"], previous["outstanding_paise"],
                format="money", tone="warning" if values["outstanding_paise"] else "neutral",
            ),
            _metric("invoices", "Fee invoices" if is_college else "Invoices", values["invoice_count"], previous["invoice_count"]),
        ],
        "series": [{
            "id": "sales_flow",
            "label": "Fees invoiced and collected" if is_college else "Billed and collected",
            "chart_type": "area",
            "format": "money",
            "grouping": "day",
            "points": _daily_series(invoices, payments, start, end, org.timezone),
        }],
        "breakdowns": [{
            "id": "invoice_status",
            "label": "Fee invoice state" if is_college else "Invoice state",
            "format": "number",
            "items": breakdown_items,
        }],
        "queues": [{
            "id": "outstanding_invoices",
            "label": "Outstanding fees" if is_college else "Outstanding invoices",
            "items": [{
                "id": row.id,
                "title": row.invoice_number,
                "detail": f"INR {(row.total_paise - row.paid_paise) / 100:,.0f} outstanding",
                "status": row.status,
                "destination": {"route": "sales", "id": row.id},
            } for row in outstanding_rows],
        }],
        "alerts": [],
        "capabilities": {
            "exports": can_export,
            "financials": True,
        },
    }


@router.get("/sales.xlsx")
def sales_excel(
    start: date | None = None,
    end: date | None = None,
    location_id: str | None = None,
    user=Depends(require_permissions("reports.view", "reports.exports")),
    _plan=Depends(require_entitlements("reports.exports")),
    db: Session = Depends(get_db),
):
    org = organization_for(db, user)
    local_today = datetime.now(ZoneInfo(org.timezone)).date()
    start = start or local_today.replace(day=1)
    end = end or local_today
    start_at, end_at = _period_bounds(org.timezone, start, end)
    rows = sales_rows(db, user, start_at, end_at, location_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = Workbook()
    sheet = book.active
    report_name = "fees" if org.industry.value == "college" else "sales"
    sheet.title = "Fees" if report_name == "fees" else "Sales"
    headers = ["Invoice", "Date", "Status", "Billed (INR)", "Paid (INR)", "Balance (INR)"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123A2F")
    for row in rows:
        sheet.append([
            row.invoice_number,
            row.created_at.astimezone(ZoneInfo(org.timezone)).isoformat(),
            row.status,
            row.total_paise / 100,
            row.paid_paise / 100,
            (row.total_paise - row.paid_paise) / 100,
        ])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = max(
            14, min(36, max(len(str(cell.value or "")) for cell in column) + 2),
        )
    output = io.BytesIO()
    book.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="edvatiq-{report_name}-{start}-{end}.xlsx"'},
    )


@router.get("/sales.pdf")
def sales_pdf(
    start: date | None = None,
    end: date | None = None,
    location_id: str | None = None,
    user=Depends(require_permissions("reports.view", "reports.exports")),
    _plan=Depends(require_entitlements("reports.exports")),
    db: Session = Depends(get_db),
):
    org = organization_for(db, user)
    local_today = datetime.now(ZoneInfo(org.timezone)).date()
    start = start or local_today.replace(day=1)
    end = end or local_today
    start_at, end_at = _period_bounds(org.timezone, start, end)
    rows = sales_rows(db, user, start_at, end_at, location_id)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    report_name = "fees" if org.industry.value == "college" else "sales"
    story = [
        Paragraph(f"{org.name} - {'Fee' if report_name == 'fees' else 'Sales'} Report", styles["Title"]),
        Paragraph(f"{start} to {end}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["Invoice", "Date", "Status", "Billed", "Paid", "Balance"]] + [[
        row.invoice_number,
        row.created_at.astimezone(ZoneInfo(org.timezone)).strftime("%d %b %Y"),
        row.status,
        f"INR {row.total_paise / 100:,.2f}",
        f"INR {row.paid_paise / 100:,.2f}",
        f"INR {(row.total_paise - row.paid_paise) / 100:,.2f}",
    ] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123A2F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), .25, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="edvatiq-{report_name}-{start}-{end}.pdf"'},
    )
